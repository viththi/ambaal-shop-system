from flask import Flask, request, redirect, url_for, session, flash, render_template_string, Response, jsonify, send_file
import mysql.connector
from mysql.connector import Error
from mysql.connector.pooling import MySQLConnectionPool
from threading import Lock
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps, lru_cache
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal, InvalidOperation
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import gzip

# ============================================================
# AMBAAL SHOP MANAGEMENT SYSTEM
# Everything is contained in this single Python file.
# ============================================================

app = Flask(__name__)

# Embedded PWA icons so the application remains a single Python file.
PWA_ICON_192 = "iVBORw0KGgoAAAANSUhEUgAAAMAAAADACAIAAADdvvtQAAAHCElEQVR42u2dy1JUVxSGVx+gEQlotxIEReQqsblKNxDxCkx8hliVWR4iE/MeGWTiPI8Q4xVFrqIixtIwsLQsMaKIF2gySUoqibFXe7rPvnzfiJKzT236fP2vtU/v08YSyTYByJeAlwAQCBAIEAgQCACBAIEAgQCBABAIEAgQCBAIAIEAgQCBAIEAEAgQCBAIEAgQCCAPSq2Y5ZmhRT8vz7nxdsNnGDP2uTBvpbFLJuMEwhu7TDJIINSxUSMjBEIdezWKWCDUsV2jyARCHTc0CrDHGSJ5VQPswSFrShjquFfOAuwhiiwQCHtcdSjAHhwyWiDscdshtnOAwQIRP86HUIA9OGSiQNjjiUP0QGCeQMSPPyFEAoFhAhE/XoUQCQQmCUT8+BZCJBAYIxDx42EIkUCAQOCAQNQvP6sYCQQIBAgEXgtEA+RtG0QCAQIBAgECAQIBIBAgECAQIBAAAgECAQIBAgEgECAQIBAgEAACAQIBAoHNlHr+98di8sNPkqjJc3g2K2e/lRfLJJCvtHXnb4+IBIGkT1HCPGZgNPozIJCtlFdI7/DnnqSuURpaEchLeo5IfFsI5xkcQyAvCevC95+QklIE8oxEjbR2hXOqympJZRDIv/Y5FjOoGUcg79ZfW0llpLIagbyh6SupqQ/zhCWl0n8CgWifWYsh0Ccpi0vfsfBP29AqdY0I5AFdQ1JRaUFfhUBetM9byZySIEAgp6lOSMfhgp08KQd7Echp0sqQWFtVxtsYAlG/tvDzj7Kxrji++2vZth2BHGVfi9QfUBz/+pVM/CJ3Z6Jf4iGQlfEze1k21mXyV92owVEEcvJPLZH0Sd2QG+dFROauyvt3ilHNKdldh0DOkUrLFzsUx68sy283RUTersmtCVPuFCBQdPVLuT6auiibm3//rKxiAyNhftSPQNGzvUo6B3RDJs9/+PnWhLx5rRibrJWWTgRyCO2mwWeP5fct/5PE+3cyd5Uq5rFA2pXRv1deUxd0Z+g7KvFyBHKC2gbZ357P+msrC9OyuqI4Q3mF9BxBIDfiR9k+P3ogj5f++Y/ZDZm+VNi2HYFMJBZT3/752J1DbRVr75GduxHIcg72qq/i5EdEuT8vL57p3M2cQiDL0a6GHizI8pP//tXmpjqEnK9ijgtUXiHdyk528nw+1e2j/fs+aWxHIGvpO6ZbS2ezMn3x/w5YuidPHxFC3gikvf2zOCsv//jEMdoq1n/c5QefXRZoV600p3RDcvnMS1vFtldJ1yAC2dk+qz7RXH8vs1c+fdjjJXn0sLCNPAJZuf66fSPXHdDaEDqUlqqdCGQVLSnZtUe5/spZC+3ujqDE2QefnRVIu/Z5uybz13I9+NkTebigbOfHEMgeyuLSd1Q3ZG5ct29Vuxbb2yz1TQhkCT1H1M/WaNuaqQsf9isW6J4CAlnTPq++lLvTuiErz+XenG5I2sUHnx0UaEdS2nt1Q2Yu6Z4ezC+0qnZKR79rr7aDt0gzI+o3+vBpGT5djLkNjsrtCadebQcTyOS7doX7ZhkECof9bbJnv8GBXyaHjyMQ8fM5MxxDIFMpKZX+k6ZPsqlDvtyLQEaSykhlFTGJQHmvcSypDhmHHnx2R6DKajmUtmOqiRpp60Ygw0iftGnjnzNVzB2B7Frd9A5LeQUCGUNdozS02DTh+DZHHnx2RCAbd9u4sUPIhc/CgkD98PLrl/L9N5LdCHMa353VfQVRa5ckauT5UxIoag72SXVSN2Tuasj2iMiM8qsXYjHJjFDC7KwFM5fDn8b8NfWeEAeqmPUCVVRK15BuyNqq7qufcy2Lr2RRucWspl6aOhAoUvqOSVlcN+TmeD7bx3JhVh9stn+2ar1AedyR0zYrutYqqxty+Lj6DYBAobG7TpoPqevXnalCzefVC7k/ry7BnYMIZE/8zF8vVP3Kuz23+mmNcAQ6Nx7Bd+DEYjKgXwYXrn791QZdUT/u09Ev1YkIrn0oV83iBGrtlGStbsjbNVmYKuysVpblwR3lNQgkbe034VksUB7rl/nrusdPi7cWG0Wg4hIvl95h4+pX3m1Q/QHZ1+K3QEVug3r02yHevZHbN4oxt+dPZWnR9BAK63rFEsm2sOZ0ZmhRwBLCEijgpQRTBIpkMQ/RXikSCEwSiBDyKn5IIDBPIELIn/ghgcBIgQghT+KHBAJTBSKEfIifwiYQDjlvT8FLGA65bQ89EBgvECHkcPwUKYFwyFV7ilfCcMhJe4raA+GQe/ZIuDsSc4SNiy69UQPn/0LscU0gHHLpVY2ghFHOXHpDRiwQGtme5UYIhEb2tgEGCYRGNnaQxgmESXatPMwVCJmsWK7aIRAYC9s5AIEAgQCBAIEAEAgQCBAIEAgAgQCBAIEAgQAQCBAIEAgQCACBAIEAgQCBwFv+BEPurkvK2inqAAAAAElFTkSuQmCC"
PWA_ICON_512 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAIAAAB7GkOtAAAUiklEQVR42u3dy3dVVZ7A8d/N0xADJDwi4WUihJiHIS9CoTyrVv0LNan/pKfdf0SvVcPqcc162F2K+KoKBESUFCpiFepCREQKeSU9wFotFlp5kb3v2Z/PyIGsnHvOPvt79rn3nlvr7NobAJSnwS4AEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAA1lyTXVAZvz04ZyewNn7/Vr+dUAG1zq699oK5HlRBADDjgx4IACZ9EAMBwLwPSiAAmPRBDAQAUz/IgABg3gclEABM/SADAoCpH2RAAMz7gBIIgKkfkIEseBic2R+cTVYAGKxgKSAAmPpBBgQAUz/IQDV5D8DsD847KwAMQbAUsALA7A/ORAHAmAPnYzW5BWSoQXW4HWQFYPYHZygCYGyB85Sf4BaQIQXV5HaQFYDZH5y5CIAxBM5fBMDoAWcxAmDcgHNZADBiwBktAMaKnQDOawEwSgBntwAYH4BzXACMDMCZLgDGBOB8FwCjAXDWC4BxADj3BcAIAMwAAuDYA+YBAQBAAGQfMBsIgOMNmBMEwJEGzAwCAIAAiDxgfhAARxcwSwiA4wqYKwQAAAFw+Q+YMUoPgNkfMG+UuwIAoLgAuPwHzB5WAAAUEwCX/4A5xAoAgGIC4PIfMJOUGACzP2A+KXcFAEBxAXD5D5hVrAAAEAAAqh0A938Ac4sVAADFBMDlP2CGsQIAoJgAuPwHzDNWAAAIAADVDoD7P4DZxgoAgGIC4PIfMOdYAQAgAAAIAAAVDIA3AAAzjxUAAAIAQLUD4P4PYP6xAgBAAAAQAAAqGABvAABmISsAAAQAAAEAQAAAqEoAvAMMmIusAAAQAAAEAAABAEAAABAAAAGoPz4DCpiRrAAAEAAABAAAAQBAAAAQAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAYNU02QXUi3/7z+jekftG/uF38b9/cKywAoDVs7u/Dmb/iJj+pWOFAMCqOvCr+tjOnt7Y3udwIQCwShqbYvxI/bTKIgABgNUyfCDaO+pmayePRUOjg4YAQHnX1B0bY3DCQUMAYMXa18fgZJ1t85S7QAgArNzksWist48rj0xHW7tDhwDAytTje6pNzfX0rjUCADl6blfs3FOf3fqVo4cAQGGX/4/0DsSWHgcQAYBlqdVi6oR6gQBQnoGx2NBVx9s/dSJqNYcRAYBlTKB1fgXdtTX2DDuMCAAsUWtbjB6q+1fhrWAEAJZs7HA0t9T9q9j/crS0OpgIACxFNZ6rXI11DAIAa6erO/qGKvJa3AVCAGApk2aFPj/TPxobNzukCAAsMgAV+gR9rRZTxx1SBAAWoW8wNm+rVs/cBUIAoLTL/0e6d8SufgcWAYCf1dwSY4cr+Lr8WDwCAP/CyMFqPkl/4mj9/aoBAgBrqqoPUFvXEUNTDi8CAD9hfWcMjFf21U17KxgBgJ8yeTwaqjskByejfb2DjADAkxw4UeVX19gUE0cdZAQA/sn2vujprfhrdBcIAYBCJ8ede+K5XQ41AgA/HIiNpdwe8TuRCAA8ZnAiOjYW8Ur9TiQCAI9Pi8VcF2/oin1jDjgCABER0dYeI9MFvV53gRAA+N7E0WhqLuj1vvSLaG1z2BEAKOn+zyMtrdV84B0CAEuzpSd6B4p71R4OigBAoTfE+4ZiU7eDjwBQsFotpk544SAAlGfPSHRttfQBAcAkWJLN26Jv0BBAAChSS2vsfznlBiws6B8CACmMHkr8cfjZU/HtzZQbMHY4mlsMBASA8hxI/fjPP/1PzJ5KuQFt7TFy0EBAACjMhk3RP5pyA+7cjvdn4vRrqSvoLhACQHGX/6kfinn2VDx8EB+ej5vXU27GwHis7zQcEACKCkDqK99H1/4LC3Hm9aRnYENMHjccEACKsas/unem3IBbX8fc2cdKkHYxBAJAKZI/CWf29Zif//6/L38Q179IuTE9vbG9z6BAAChAY1OMp/71x5nHr/rPnExdRD8WjwBQgqGpaO9IuQE3rsXHFx7vwauJ98nE0WhoNDQQAKou+dXumZM//g7w3z6KL/6acpM6NsbghKGBAFBp7etjcDLxNjzxej/5W8FTvhCAAFBtE0ejsSnlBly7Gp9eelIAUt8FGpmOdc8aIAgA1ZX84/8/dbv/i7/G1Y9TblhTc4wfMUAQACrquV2xa2/ibfiZWz0z7gIhAFDVy/+rl+PzKz/dhtR3gXoHYut2wwQBoHJqtZhK/cyDn3+n9/oX8clc6Y1EAGD17RuLDZsSb8O//Lx/8kXA5PHEz8hDAOApXNumfuLNJ3Nx/fN/FYCTiX8mrGtr7BkxWBAAKqS1LV46lHgbFnN1f/N6fPRe6lK6C4QAUCVjh6OlNeUGLCzE6cU98Cf5Z4H2vxwtzxgyCABVkfz+z+J/+OWHDwpNtVoaPWTIIABUQtfWeGE48TYs/kkP3978/58KSNZLd4EQACpy+f/LxJ9smX+4tJ/9Sv5coP7R2LjZwEEAqH9Tqe//XDwbt79Zwv//6OeCE8rhOxMIAKxU32Bs6Um8DTN/XNr/f+d2vD+TetnkJ2IQAOpd8tvZD+7Hu28t+V8lvwvUvSN29xs+CAB1q6k5xg4n3oYLf447t5f8r959O+7fK72dCAAs38jBaGtPvA3Lu5a/eyfeeyfxlif/7QQEAJZvOvU17L3v4vzba1qOVbSuI4YPGEQIAHWoY2MMjCfehnffjnt3l/lv3/tT3L2TePvdBUIAqEuTx6OhMfE2rOTpnvfvxbm3Em//4GS0rzeUEADqTfL7P3dux4WVfZoz+dOhG5ti4qihhABQV7b3Rk9v4m1Y+fe5PjgTf7+VuqO+EIAAUF9y+B7Tyt/Fffggzr6R+FXs3BPbdhtQCAD1MqoaY/JY4m249fXqPNMt+dOhI4NnaSAAsFgvjkfHxsTbMHtqdZ7qfOlc3Po6fQD8TiQCQH3I4f7PzCq9fzs/H7OvJ34tG7pi35hhhQCQvbb29F9funEtPr6wei3J4C6QLwQgANSB8SPR3JJ4G86s6m+7f3whblxL/IpGD0Vrm8GFAJC3HK5VV/cpDgsLceZk4lfU3JL+sXoIAPycLT3R+2Libbh2Na78JeuiLM+0u0AIAC7/136yvvKX+PKzxK+rbyg2dRtiCABZyuSHDJ/S1XryRUCt5gsBCAC52jMcXakvUa9ejs8+qWYAwmeBEACyNVXR+z9POy2Lt3lb9A0aaAgAmWlpjbFX0m/GzNN8fmfyzwJZBCAA5CiHD6pfmYvrn9drXRZp7HD6r1kgAPCYHO7/PO0J+trV+PTDxK+xrT1GDhpuCADZ2LAp9u1PvA0LC3H66d+iOZ3BIsBdIASAnC7/j6d/XOVH78XN608/AK+t5kMmlmdgPNZ3GnQIAHnI4Zp0bR7ZduNaXL6Y+qRtiMnjBh0CQAZ29cdzuxJvw/zDtXtos7tACAD8YzLK4OupF8/GtzfX6G+t7qNGl6fn+djeZ+ghACTV2BQTR9NvxlpelX9zIy69m/4l+7F4BIDEhqaifX3ibXhwP869uaZ/MYcvBEwcjYZGAxABIJ0cbkZf+HPcub2mf/HsqXj4IPGr7tgYgxMGIAJAIu0dMTSVfjPW/jFtt2/FxVn1RQAo2MSxaGxKvA33vovzbxdRnX82PB3rnjUMEQBKvQI9/07cu5vg7557Mx7cT/zam5pj/IhhiACw5rp3xq696Tdj5o9p/u53f48Lf9ZgBIAi5fAxxDu348JMsr+ew12g5wdi63aDEQFgDdVqMXks/WacfSPlp3HOvx33vrMIQAAozL79sXFz+s1Iew1+726cfyf9Tpg6kf5JfAgABcnhqvPW1zE3m3gbcvhGWOeW2DNiSCIArInWtnjpUPrNmD0V8/OJt+H9mbX+DtoTeSwEAsAaGXslWlrTb0YOT+V8cD/efSv9ZoweipZnDEwEgKcvh19//PrL+OhCFnsjh7tArW0xesjARAB4yrq2xp7hDC7/M/hlrkcuzsbtb9Jvhs8CIQCsxUSTw2dOcvgM/iPzD2P2VPrN6B/N4nNZCABVNpXBz79cuxpX/pLRPsnh3YhaLYtDgwBQWb0vxpae9Jtx5mReu+XS+bj5VRaLM1iSJruAuptifv2b+PVvHI0f694Ru/vjkzl7AisAVv1iwbMn8y+0LwQgADwNIwejrd1uyNrEkfS/0IAAUMWrS7eYs7euI4YP2A0IAKuqY2O8OG436DQCQHkmj0VDo91QBwYno3293YAAsIrXld5drBONTVn8VAMCQEX09Mb2XruhfmrtLhACwGqZNqHUlZ17YttuuwEBYOVDpCEmjtkNdcZjIRAAVsHARKzvtBvqLwANTm4EgBVy/6cebeiKffvtBgSAFWhrj+Fpu6E+FwHKjQCwEmOHo7nFbqhLo4fimXV2AwLAcvm18frV3BJjr9gNCADLsnlb9L5oN9QxXwhAADB9FKpvKDZ12w0IAEvkJwYdRASAQr0w7OLRMg4BwMRB3dq8LfqG7AYEgEVrafUBkgq13F0gBIDFe+lQtLbZDRXhyxwIAK4ZC9XWHiMH7QYEgEXYsCn2jdkN1Sq6d3QQABZj6njUanZDpQyMe6QrAsBiAuD+T/XO84aYPG43IAD8rF17/ZhUNbkLhABgmihUz/Ox4wW7AQHgJzQ2xcRRu0HdEQDKMzQV7evthsqaOBoNjXYDAoArxPJ0bIzBCbsBAeCfrOuIoSm7QeMRAIq8P9DYZDdU3PB0rHvWbuB7zni+N53HteG5N+N3/17B3dvYFP/xX+kn36bmGD8Sr/+38Y4VAP/QvSN29WexJadfq+Yefvggzr2RxZa4C4QA8PikkMePv9/7Ls6/XdmdfPpkFpvx/EBs3W7IIwBExKMfDszjOQHn34l7dyu7n+fOxrc3LQIQAHLSPxobN2exJTOvVnk/zz+M2VNZbMnUCc/7QwB4dD2Yx/2fO7fj/ZmK7+ozebzD0bkl9owY+AhA8VrbYvRQFlty7s14cL/ie/vS+fjmqyy2ZPpXxj4CULz9L0dLaxZbcvrV6u/thYVc7gKNHoqWZwx/AaBsmdz/+fZmXDxbxA7P5H2OfFZ+CABpdG6JPcNZbMns6zH/sIh9fvmDuHEtiy2Z9lkgAbALir78/2UunwaZea2Ufb6wEGfy+ELA3tHo3OIkEABKlcmvP968Hh+9V9Buz+QbYbWa34kUAErVm803Qk+fjIWFgvb8lbn48rNcloAIACU6kM0HAUv4/M+PZHIXqHtH7O53KghAvfn9W4btijQ1x9jhLLbk+ufxyVxx+z+fZ94d8IWAgmckK4BCjWTzXPjTr5W4///2cXzx1yy2ZOKI34GwAqAwU9nc/J15tdBDkEn51nXE8AEnhABQjHx+G/bzK3H1sgAk5i6QAFCQyWPR0GgSTOyLT3OJ3+BEPLvBaSEAlOGA+z8WAT/Q2BQTR50WAkABep6P7X1ZbMmnH8a1qwKQxzWBu0ACUF98ErTeL/9Pv1r6sfjys/j0UhZbsvOF2LbbyVHcXGQFUFjwG3L59v/CQtFvAOS4CPCtYCsAqm1gItZ3ZrEl+TwUM3kAMnkMxuTxaDAfCAAVduBELlsy86qjERFx41pcvpjFlmzoin37HRABoKLa2mPkYBZbMj+fy8NwcnDGW8EIwDJ4H3hJxg5Hc0sWW3LpXNz62gH5Xj4PQ33pF/HMOgekoFnICqAgGX3839u/P/DNV/Hh+Sy2pLklxl5xQKwAqJzN26JvMIstefggzr7hgPx4EeAqAQGg+if2B2fi77cckMfMvh7z81lsSd9QbOp2QASgTngbYDFqtVx+/TF8/+tJvr0Zc2dzGSoWAeXMP1YARcjnsu7+vTj3lgPyBPl8FiifawUEgFUwnc013Xt/irt3HJAnOPtGPHyQxZZs3hZ9Qw5IGfcGOrv2VuBl/PbgnGMJrKUK3H+2AgAolAAACAAAAlB3fBgUMOdYAQBQWAAsAgCzjRUAAAIAQAkBcBcIMM9YAQBQWAAsAgAzjBUAAIUFwCIAMLdYAQAgAACUEAB3gQCzihUAAIUFwCIAMJ+UuwLQAMBMUmgAACg3ABYBgDnECgCAwgJgEQCYPawAACgsABYBgHmj3BWABgBmjEIDAEC5AbAIAMwV5a4ANAAwSxQaAA0AzA/lBgCAcgNgEQCYGcpdAWgAYE5ocLwBs3+ZL9x7AACFKjoAFgFAyfNAg2PvBAAzgAAYAYBzXwCMA8BZLwBGA+B8FwBjAnCmC4CRATjHBcD4AJzdAmCUAM5rATBWAGe0ABgxgHNZAIwbwFksAEYP4PwVAGMIcOauvVpn1157YZF+e3DOTgBTvxWAUQU4TwXA2AKcoXXILaBlcjsITP1WAEYb4HwUAGMOcCbWD7eAVoHbQWDqtwIwCgHnnRWApQBg6hcAGQBM/QIgA4CpPxfeAzBewdlkBYClAJj6BQAlAPO+ACADYOoXAGQATP0CgBKAeV8AkAEw9QsAYgAmfQFACcC8LwCIASZ9BAA9wIyPAKAKmOsRAAAS8DA4AAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAGwCwAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAABI6v8AC55d01JbIy4AAAAASUVORK5CYII="

# -----------------------------------------------------------------
# APPLICATION / SECURITY SETTINGS
# -----------------------------------------------------------------
# On Render, add SECRET_KEY as an Environment Variable.
# For local development, the fallback below is used.
app.secret_key = os.environ.get(
    "SECRET_KEY",
    "ambaal-local-development-secret-change-before-production"
)

# Keep sessions reasonably secure when HTTPS is used in production.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"
)

# Compress large HTML/JSON/JS responses before sending them to phones/browsers.
# This reduces transfer time for the large single-file interface.
@app.after_request
def compress_large_responses(response):
    try:
        accepted = request.headers.get("Accept-Encoding", "")
        content_type = (response.content_type or "").lower()
        compressible = any(t in content_type for t in ("text/html", "application/json", "javascript", "text/css"))
        if (
            "gzip" in accepted.lower()
            and compressible
            and not response.direct_passthrough
            and not response.headers.get("Content-Encoding")
            and 200 <= response.status_code < 300
        ):
            data = response.get_data()
            if len(data) >= 1500:
                compressed = gzip.compress(data, compresslevel=5)
                if len(compressed) < len(data):
                    response.set_data(compressed)
                    response.headers["Content-Encoding"] = "gzip"
                    response.headers["Content-Length"] = str(len(compressed))
                    response.headers["Vary"] = "Accept-Encoding"
    except Exception:
        pass
    return response

# ------------------------- DATABASE --------------------------
# LOCAL XAMPP defaults:
#   DB_HOST=localhost
#   DB_PORT=3306
#   DB_USER=root
#   DB_PASSWORD=
#   DB_NAME=ambaal_shop
#
# RENDER / ONLINE MYSQL:
# Add DB_HOST, DB_PORT, DB_USER, DB_PASSWORD and DB_NAME
# as Environment Variables in Render.
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "3306"))
DB_USER = os.getenv("DB_USER", "root")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_NAME = os.getenv("DB_NAME", "ambaal_shop")

# Performance settings.
# FAST_START avoids re-running all CREATE/ALTER checks on every Gunicorn boot
# once the expected schema already exists.
FAST_START = os.getenv("FAST_START", "true").lower() == "true"

# Page-view logging is useful for auditing but expensive on hosted databases
# because the old code opened a second MySQL connection on every GET request.
# Important actions (POST requests, login and logout) are still logged.
LOG_PAGE_VIEWS = os.getenv("LOG_PAGE_VIEWS", "false").lower() == "true"

# If the database already exists in production, set this to true on Render to
# avoid a database connection during web-worker boot. Keep false for a brand-new
# database so tables can be created automatically.
SKIP_DB_INIT_ON_STARTUP = os.getenv("SKIP_DB_INIT_ON_STARTUP", "false").lower() == "true"


def server_connection():
    """Connect to MySQL without selecting a database."""
    return mysql.connector.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        connection_timeout=10
    )


# Reuse authenticated MySQL connections instead of opening a new remote
# connection for every page click. This is one of the biggest performance
# improvements when Flask and MySQL are hosted on different servers.
DB_POOL_SIZE = max(1, min(int(os.getenv("DB_POOL_SIZE", "5")), 20))
_db_pool = None
_db_pool_lock = Lock()


def _get_db_pool():
    global _db_pool
    if _db_pool is None:
        with _db_pool_lock:
            if _db_pool is None:
                _db_pool = MySQLConnectionPool(
                    pool_name=f"ambaal_{os.getpid()}",
                    pool_size=DB_POOL_SIZE,
                    pool_reset_session=False,
                    host=DB_HOST,
                    port=DB_PORT,
                    user=DB_USER,
                    password=DB_PASSWORD,
                    database=DB_NAME,
                    connection_timeout=10
                )
    return _db_pool


def db_connection():
    """Borrow a reusable MySQL connection from the per-worker connection pool."""
    connection = _get_db_pool().get_connection()
    # Keep TIMESTAMP/NOW() behaviour in Sri Lanka time. The session setting is
    # retained because pooled sessions are not reset when returned to the pool.
    if not getattr(connection, "_ambaal_timezone_ready", False):
        cursor = connection.cursor()
        cursor.execute("SET time_zone = '+05:30'")
        cursor.close()
        try:
            connection._ambaal_timezone_ready = True
        except Exception:
            pass
    return connection


def database_schema_ready():
    """Return True when the current database already has the required schema."""
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT
                COUNT(DISTINCT table_name) AS table_count,
                SUM(table_name = 'shop_items' AND column_name = 'finished_at') AS has_finished_at,
                SUM(table_name = 'prices' AND column_name = 'buying_price') AS has_buying_price
            FROM information_schema.columns
            WHERE table_schema = %s
              AND table_name IN ('users','customers','loan_transactions','shop_items','prices','activity_logs')
            """,
            (DB_NAME,)
        )
        row = cursor.fetchone() or (0, 0, 0)
        return int(row[0] or 0) == 6 and int(row[1] or 0) >= 1 and int(row[2] or 0) >= 1
    except Exception:
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


def ensure_performance_indexes():
    """Create only missing indexes used by the busiest pages/searches.

    The check is one metadata query. Existing indexes are left untouched, so
    normal starts remain quick after the first successful setup.
    """
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT table_name, index_name
            FROM information_schema.statistics
            WHERE table_schema = %s
              AND table_name IN ('customers','loan_transactions','shop_items','prices','activity_logs')
            """,
            (DB_NAME,)
        )
        existing = {(row[0], row[1]) for row in cursor.fetchall()}

        wanted = [
            ('customers', 'idx_customers_name', 'customer_name'),
            ('customers', 'idx_customers_mobile', 'mobile_number'),
            ('loan_transactions', 'idx_loan_customer_date', 'customer_id, transaction_date'),
            ('loan_transactions', 'idx_loan_date_id', 'transaction_date, id'),
            ('shop_items', 'idx_shop_items_name', 'item_name'),
            ('shop_items', 'idx_shop_items_finished', 'finished_at'),
            ('prices', 'idx_prices_item_name', 'item_name'),
            ('prices', 'idx_prices_updated', 'updated_at'),
        ]

        changed = False
        for table_name, index_name, columns in wanted:
            if (table_name, index_name) not in existing:
                cursor.execute(f"CREATE INDEX `{index_name}` ON `{table_name}` ({columns})")
                changed = True
        if changed:
            connection.commit()
    except Exception as error:
        print(f"Performance index setup skipped: {error}")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


def initialize_database():
    """
    Prepare the MySQL database and application tables.

    This works in two common situations:
    1. Local XAMPP/MySQL, where the application can create DB_NAME.
    2. Hosted MySQL, where DB_NAME already exists and CREATE DATABASE
       permission may not be available.
    """

    # First try to create the database. Some hosted MySQL providers do not
    # allow CREATE DATABASE, so failure here is not automatically fatal.
    server = None
    cursor = None
    try:
        server = server_connection()
        cursor = server.cursor()
        cursor.execute(
            f"CREATE DATABASE IF NOT EXISTS `{DB_NAME}` "
            "CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
        )
        server.commit()
    except Error as create_error:
        print(
            "Database creation step skipped or unavailable: "
            f"{create_error}"
        )
    finally:
        if cursor:
            cursor.close()
        if server and server.is_connected():
            server.close()

    # DB_NAME must now exist (either created above or supplied by the host).
    connection = db_connection()
    cursor = connection.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(100) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INT AUTO_INCREMENT PRIMARY KEY,
            customer_code VARCHAR(50) NOT NULL UNIQUE,
            customer_name VARCHAR(150) NOT NULL,
            mobile_number VARCHAR(30) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS loan_transactions (
            id INT AUTO_INCREMENT PRIMARY KEY,
            customer_id INT NOT NULL,
            transaction_date DATE NOT NULL,
            transaction_type ENUM('LOAN', 'PAYMENT') NOT NULL,
            amount DECIMAL(12,2) NOT NULL,
            note VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT fk_loan_customer
                FOREIGN KEY (customer_id) REFERENCES customers(id)
                ON DELETE CASCADE
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shop_items (
            id INT AUTO_INCREMENT PRIMARY KEY,
            item_name VARCHAR(150) NOT NULL,
            item_code VARCHAR(50) UNIQUE,
            quantity INT NOT NULL DEFAULT 0,
            description VARCHAR(255),
            finished_at DATETIME NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            id INT AUTO_INCREMENT PRIMARY KEY,
            item_name VARCHAR(150) NOT NULL,
            buying_price DECIMAL(12,2) NOT NULL DEFAULT 0.00,
            selling_price DECIMAL(12,2) NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP
        )
    """)

    # Backward-compatible upgrades for databases created by older versions.
    cursor.execute("SHOW COLUMNS FROM shop_items LIKE 'finished_at'")
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE shop_items ADD COLUMN finished_at DATETIME NULL AFTER description")

    cursor.execute("SHOW COLUMNS FROM prices LIKE 'buying_price'")
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE prices ADD COLUMN buying_price DECIMAL(12,2) NOT NULL DEFAULT 0.00 AFTER item_name")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS activity_logs (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NULL,
            username VARCHAR(100) NULL,
            activity VARCHAR(180) NOT NULL,
            endpoint VARCHAR(120) NULL,
            method VARCHAR(10) NULL,
            ip_address VARCHAR(64) NULL,
            device_type VARCHAR(60) NULL,
            browser VARCHAR(60) NULL,
            platform VARCHAR(60) NULL,
            user_agent VARCHAR(500) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_activity_created (created_at),
            INDEX idx_activity_user (username),
            CONSTRAINT fk_activity_user
                FOREIGN KEY (user_id) REFERENCES users(id)
                ON DELETE SET NULL
        )
    """)

    # Keep the live log table compact. Older entries are not needed for
    # day-to-day device monitoring.
    cursor.execute(
        "DELETE FROM activity_logs WHERE created_at < (NOW() - INTERVAL 30 DAY)"
    )

    # Create the first admin only when no "ambaal" user exists.
    # IMPORTANT: set ADMIN_PASSWORD in Render before first deployment.
    default_admin_username = os.getenv("ADMIN_USERNAME", "ambaal")
    default_admin_password = os.getenv("ADMIN_PASSWORD", "ambaal")

    cursor.execute(
        "SELECT id FROM users WHERE username = %s",
        (default_admin_username,)
    )
    admin = cursor.fetchone()

    if not admin:
        cursor.execute(
            "INSERT INTO users (username, password_hash) VALUES (%s, %s)",
            (
                default_admin_username,
                generate_password_hash(default_admin_password)
            )
        )
        print(
            f"Default administrator '{default_admin_username}' created."
        )

    connection.commit()
    cursor.close()
    connection.close()


# ----------------------- AUTHENTICATION -----------------------

def login_required(view_function):
    @wraps(view_function)
    def wrapped_view(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        return view_function(*args, **kwargs)
    return wrapped_view


# ------------------------- UTILITIES --------------------------

def money(value):
    try:
        return f"{Decimal(value):,.2f}"
    except Exception:
        return "0.00"


app.jinja_env.filters["money"] = money


# All application database sessions use Sri Lanka Standard Time (UTC+05:30).
SRI_LANKA_TZ = ZoneInfo("Asia/Colombo")


def to_sri_lanka_time(value):
    """Format a database datetime that is already returned in Sri Lanka time."""
    if not isinstance(value, datetime):
        return value or "-"
    return value.strftime("%Y-%m-%d %H:%M:%S")


app.jinja_env.filters["sl_time"] = to_sri_lanka_time


def get_customer_balance(connection, customer_id):
    cursor = connection.cursor(dictionary=True)
    cursor.execute("""
        SELECT COALESCE(SUM(
            CASE
                WHEN transaction_type = 'LOAN' THEN amount
                WHEN transaction_type = 'PAYMENT' THEN -amount
                ELSE 0
            END
        ), 0) AS balance
        FROM loan_transactions
        WHERE customer_id = %s
    """, (customer_id,))
    result = cursor.fetchone()
    cursor.close()
    return Decimal(result["balance"] or 0)


def client_ip_address():
    """Return the best available client IP when running locally or behind Render."""
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()[:64]
    return (request.remote_addr or "Unknown")[:64]


def detect_client_device(user_agent):
    """Return broad device, browser and platform labels from a User-Agent string."""
    ua = (user_agent or "").lower()

    if "iphone" in ua:
        device, platform = "iPhone", "iOS"
    elif "ipad" in ua:
        device, platform = "iPad", "iPadOS"
    elif "android" in ua:
        device, platform = ("Android Phone" if "mobile" in ua else "Android Tablet"), "Android"
    elif "macintosh" in ua or "mac os x" in ua:
        device, platform = "Mac", "macOS"
    elif "windows" in ua:
        device, platform = "Windows PC", "Windows"
    elif "linux" in ua:
        device, platform = "Computer", "Linux"
    else:
        device, platform = "Unknown Device", "Unknown"

    if "edg/" in ua:
        browser = "Microsoft Edge"
    elif "crios/" in ua:
        browser = "Chrome"
    elif "chrome/" in ua and "edg/" not in ua:
        browser = "Chrome"
    elif "fxios/" in ua or "firefox/" in ua:
        browser = "Firefox"
    elif "safari/" in ua and "chrome/" not in ua and "crios/" not in ua:
        browser = "Safari"
    else:
        browser = "Other Browser"

    return device, browser, platform


def log_activity(activity, user_id=None, username=None):
    """Write one activity event without allowing logging failures to break the app."""
    connection = None
    cursor = None
    try:
        ua = request.headers.get("User-Agent", "")[:500]
        device, browser, platform = detect_client_device(ua)
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO activity_logs
                (user_id, username, activity, endpoint, method, ip_address,
                 device_type, browser, platform, user_agent)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id if user_id is not None else session.get("user_id"),
            username if username is not None else session.get("username"),
            activity[:180],
            (request.endpoint or "")[:120],
            request.method[:10],
            client_ip_address(),
            device, browser, platform, ua
        ))
        connection.commit()
    except Exception as error:
        print(f"Activity log skipped: {error}")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


def activity_label():
    """Convert Flask endpoint names into readable activity text."""
    labels = {
        "dashboard": "Viewed Dashboard",
        "customer_loans": "Viewed Customer Loans",
        "customer_details": "Viewed Customer Details",
        "add_customer": "Opened / Added Customer",
        "add_transaction": "Added Loan or Payment",
        "edit_transaction": "Edited Transaction",
        "delete_transaction": "Deleted Transaction",
        "delete_customer": "Deleted Customer",
        "shop_items": "Viewed / Updated Shop Items",
        "delete_item": "Deleted Shop Item",
        "finish_item": "Marked Shop Item Finished",
        "reopen_item": "Reopened Shop Item",
        "price_management": "Viewed / Updated Prices",
        "delete_price": "Deleted Price",
        "all_data": "Viewed Data Store",
        "download_excel": "Downloaded Excel File",
        "live_logs": "Viewed Live Logs",
    }
    return labels.get(request.endpoint, f"Opened {request.path}")


@app.before_request
def record_authenticated_activity():
    """Record important activity without slowing normal page navigation."""
    if "user_id" not in session:
        return
    if request.endpoint in {"pwa_manifest", "pwa_icon", "service_worker", "live_logs_data"}:
        return

    # GET page-view logging caused an additional remote DB connection + INSERT
    # before every page could load. Keep it optional for deployments that need it.
    if request.method == "GET" and not LOG_PAGE_VIEWS:
        return

    log_activity(activity_label())


# -------------------------- DESIGN ----------------------------

BASE_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
    <meta name="theme-color" content="#111225">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="Ambaal Shop">
    <link rel="manifest" href="{{ url_for('pwa_manifest') }}">
    <link rel="icon" sizes="192x192" href="{{ url_for('pwa_icon', size=192) }}">
    <link rel="apple-touch-icon" href="{{ url_for('pwa_icon', size=192) }}">
    <title>{{ title }} | Ambaal Shop</title>
    <style>
        :root {
            --primary: #6d4aff;
            --primary-2: #8b5cf6;
            --primary-dark: #5433e8;
            --sidebar: #111225;
            --sidebar-soft: #1b1d38;
            --background: #f5f7fb;
            --card: rgba(255,255,255,.96);
            --text: #202235;
            --muted: #747b92;
            --success: #15945f;
            --danger: #d94343;
            --warning: #b87300;
            --border: #e6e9f2;
            --shadow: 0 12px 35px rgba(33, 38, 73, .08);
            --radius: 18px;
        }

        * { box-sizing: border-box; }
        html { scroll-behavior: smooth; }

        body {
            margin: 0;
            font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background:
                radial-gradient(circle at 10% 0%, rgba(109,74,255,.08), transparent 28%),
                radial-gradient(circle at 90% 8%, rgba(58,180,255,.08), transparent 24%),
                var(--background);
            color: var(--text);
            min-height: 100vh;
        }

        a { text-decoration: none; color: inherit; }
        button, input, select, textarea { font: inherit; }

        .layout { min-height: 100vh; display: flex; }

        .sidebar {
            width: 272px;
            background:
                radial-gradient(circle at 20% 10%, rgba(139,92,246,.22), transparent 25%),
                linear-gradient(180deg, #111225 0%, #15172d 60%, #111225 100%);
            color: white;
            padding: 24px 16px;
            position: fixed;
            inset: 0 auto 0 0;
            overflow-y: auto;
            z-index: 90;
            border-right: 1px solid rgba(255,255,255,.05);
            box-shadow: 18px 0 45px rgba(16,18,37,.08);
        }

        .brand {
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 22px;
            font-weight: 900;
            letter-spacing: .2px;
            padding: 5px 10px 24px;
        }

        .brand-logo {
            width: 42px;
            height: 42px;
            border-radius: 13px;
            display: grid;
            place-items: center;
            font-weight: 900;
            color: white;
            background: linear-gradient(135deg, var(--primary), #a78bfa);
            box-shadow: 0 10px 25px rgba(109,74,255,.35);
        }

        .brand span { color: #a996ff; }

        .nav-label {
            margin: 18px 12px 8px;
            color: #858aa4;
            font-size: 11px;
            font-weight: 800;
            text-transform: uppercase;
            letter-spacing: 1.2px;
        }

        .nav-link {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 12px 13px;
            margin: 6px 0;
            border-radius: 12px;
            color: #d9dbea;
            font-weight: 650;
            transition: all .2s ease;
            position: relative;
        }

        .nav-link svg { width: 19px; height: 19px; stroke: currentColor; flex: 0 0 19px; }
        .nav-link:hover { background: rgba(255,255,255,.07); color: white; transform: translateX(2px); }
        .nav-link.active {
            background: linear-gradient(135deg, var(--primary), #7047ff 55%, #815cf7);
            color: white;
            box-shadow: 0 10px 24px rgba(109,74,255,.28);
        }

        .sidebar-footer {
            margin-top: 28px;
            padding: 14px;
            border: 1px solid rgba(255,255,255,.08);
            background: rgba(255,255,255,.04);
            border-radius: 14px;
            color: #a9aec4;
            font-size: 12px;
            line-height: 1.5;
        }

        .main { margin-left: 272px; width: calc(100% - 272px); min-height: 100vh; }

        .topbar {
            min-height: 76px;
            background: rgba(255,255,255,.82);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-bottom: 1px solid rgba(225,228,238,.9);
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 0 28px;
            position: sticky;
            top: 0;
            z-index: 60;
        }

        .topbar h1 { font-size: 22px; margin: 0; letter-spacing: -.35px; }
        .topbar-left { display:flex; align-items:center; gap:12px; min-width:0; }

        .user-area { display: flex; align-items: center; gap: 12px; color: var(--muted); }
        .admin-pill {
            display:flex; align-items:center; gap:9px;
            padding: 8px 11px 8px 8px;
            border-radius: 999px;
            background: white;
            border: 1px solid var(--border);
            box-shadow: 0 5px 16px rgba(31,35,64,.05);
        }
        .avatar {
            width: 32px; height: 32px; border-radius: 50%; display:grid; place-items:center;
            background: linear-gradient(135deg, var(--primary), #9b8afc);
            color:white; font-size:13px; font-weight:900;
        }
        .logout {
            display:inline-flex; align-items:center; gap:7px;
            color: var(--danger); font-weight: 800;
            padding: 9px 11px; border-radius:10px;
            transition:.2s;
        }
        .logout:hover { background:#fff0f0; }
        .logout svg { width:17px; height:17px; stroke:currentColor; }

        .content { padding: 28px; max-width: 1600px; margin: 0 auto; }

        .grid { display: grid; grid-template-columns: repeat(3,minmax(0,1fr)); gap: 18px; }

        .card {
            background: var(--card);
            border: 1px solid rgba(227,230,239,.95);
            border-radius: var(--radius);
            padding: 22px;
            box-shadow: var(--shadow);
            margin-bottom: 20px;
            transition: box-shadow .2s ease, transform .2s ease;
        }
        .card:hover { box-shadow: 0 16px 42px rgba(33,38,73,.10); }
        .card h2, .card h3 { margin-top: 0; letter-spacing: -.25px; }

        .stat-title { color: var(--muted); font-size: 13px; font-weight: 700; }
        .stat-value { font-size: 31px; font-weight: 900; margin-top: 8px; letter-spacing: -.8px; }

        .section-header { display:flex; align-items:center; justify-content:space-between; gap:15px; margin-bottom:18px; }
        .section-header h2 { margin:0; font-size:22px; }

        .form-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px; }
        .form-group { margin-bottom:4px; }
        label { display:block; margin-bottom:7px; font-size:13px; font-weight:800; color:#4c5269; }

        input, select, textarea {
            width:100%; padding:12px 13px; border:1px solid #d9deea; border-radius:11px;
            font-size:15px; background:#fff; color:var(--text); transition:.2s ease;
        }
        input:hover, select:hover, textarea:hover { border-color:#c7cde0; }
        input:focus, select:focus, textarea:focus {
            outline:none; border-color:var(--primary); box-shadow:0 0 0 4px rgba(109,74,255,.11);
        }
        textarea { min-height:90px; resize:vertical; }
        .full { grid-column:1 / -1; }

        .btn {
            display:inline-flex; align-items:center; justify-content:center; gap:7px;
            border:0; border-radius:10px; padding:11px 16px; font-size:14px; font-weight:800;
            cursor:pointer; white-space:nowrap; transition:transform .18s ease, box-shadow .18s ease, background .18s ease;
        }
        .btn:hover { transform: translateY(-1px); }
        .btn svg { width:16px; height:16px; flex:0 0 16px; stroke:currentColor; }
        .btn-primary { background:linear-gradient(135deg,var(--primary),#7453ff); color:#fff; box-shadow:0 8px 18px rgba(109,74,255,.22); }
        .btn-primary:hover { background:linear-gradient(135deg,var(--primary-dark),#6844ff); box-shadow:0 10px 22px rgba(109,74,255,.29); }
        .btn-secondary { background:#f0edff; color:var(--primary-dark); }
        .btn-secondary:hover { background:#e7e2ff; }
        .btn-danger { background:#fff0f0; color:var(--danger); }
        .btn-danger:hover { background:#ffe5e5; }
        .btn-success { background:#e9f9f1; color:var(--success); }
        .btn-success:hover { background:#dcf5e8; }

        .action-buttons { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
        .action-buttons form { margin:0; }

        .table-wrap { overflow-x:auto; border-radius:14px; border:1px solid #edf0f5; }
        table { width:100%; border-collapse:separate; border-spacing:0; min-width:760px; background:white; }
        th,td { text-align:left; padding:14px 13px; border-bottom:1px solid #eef0f5; font-size:14px; }
        th { color:#71788f; background:#f8f9fd; font-size:11px; text-transform:uppercase; letter-spacing:.5px; font-weight:900; }
        th:first-child { border-top-left-radius:13px; }
        th:last-child { border-top-right-radius:13px; }
        tr:last-child td { border-bottom:0; }
        tbody tr { transition:background .16s ease; }
        tbody tr:hover td { background:#fbfbff; }

        .badge { display:inline-flex; align-items:center; border-radius:999px; padding:5px 9px; font-size:11px; font-weight:900; }
        .badge-loan { background:#fff1e6; color:#a9540c; }
        .badge-payment { background:#e7f8ef; color:var(--success); }
        .balance-positive { color:var(--danger); font-weight:900; }
        .balance-zero { color:var(--success); font-weight:900; }

        .alert { border-radius:12px; padding:13px 15px; margin-bottom:18px; font-weight:700; border:1px solid transparent; box-shadow:0 7px 18px rgba(31,35,64,.05); }
        .alert-success { background:#e9f9f1; color:#137a4f; border-color:#cef0df; }
        .alert-danger { background:#fff0f0; color:#bb3333; border-color:#ffdada; }
        .alert-warning { background:#fff7e4; color:#9b6200; border-color:#f5e3b6; }
        .alert-info { background:#edf5ff; color:#245f98; border-color:#d9e9fb; }

        .empty { text-align:center; padding:40px 20px; color:var(--muted); background:linear-gradient(180deg,#fff,#fbfcff); border-radius:13px; }

        .search-row { display:flex; gap:10px; margin-bottom:18px; flex-wrap:wrap; align-items:center; }
        .search-row input { max-width:460px; min-width:220px; flex:1 1 340px; background:#fbfcff; }

        .login-page {
            min-height:100vh; display:grid; place-items:center; padding:20px;
            background:
                radial-gradient(circle at 12% 15%, rgba(139,92,246,.40), transparent 25%),
                radial-gradient(circle at 90% 10%, rgba(53,183,255,.22), transparent 20%),
                linear-gradient(135deg,#111225,#24264a 60%,#17182d);
        }
        .login-card { width:100%; max-width:430px; background:rgba(255,255,255,.97); border-radius:24px; padding:35px; box-shadow:0 25px 70px rgba(0,0,0,.28); border:1px solid rgba(255,255,255,.35); }
        .login-card h1 { margin-bottom:8px; font-size:29px; }
        .login-card p { color:var(--muted); margin-top:0; }
        .login-card .form-group { margin:18px 0; }
        .login-card .btn { width:100%; padding:13px; }

        .install-app-btn {
            display:none; align-items:center; gap:7px; border:1px solid #e6e1ff;
            background:#f4f1ff; color:var(--primary-dark); border-radius:11px;
            padding:9px 12px; font-weight:850; cursor:pointer; transition:.2s ease;
        }
        .install-app-btn.show { display:inline-flex; }
        .install-app-btn:hover { background:#ebe5ff; transform:translateY(-1px); }
        .install-app-btn svg { width:17px; height:17px; stroke:currentColor; }

        .network-banner {
            display:none; position:fixed; left:50%; bottom:24px; transform:translateX(-50%);
            z-index:200; padding:10px 14px; border-radius:999px; font-size:13px; font-weight:850;
            box-shadow:0 12px 35px rgba(15,18,38,.18);
        }
        .network-banner.offline { display:block; background:#fff1f1; color:#b93030; border:1px solid #ffd8d8; }
        .network-banner.online { display:block; background:#eafaf1; color:#137c50; border:1px solid #ccefdc; }

        .bottom-nav { display:none; }

        @media (display-mode: standalone) {
            .install-app-btn { display:none !important; }
        }

        .mobile-menu { display:none; }
        .menu-overlay { display:none; }

        @media (max-width: 1100px) {
            .content { padding:22px; }
            .grid { gap:14px; }
        }

        @media (max-width: 900px) {
            .sidebar { transform:translateX(-100%); transition:.25s ease; width:min(82vw,290px); }
            .sidebar.open { transform:translateX(0); }
            .main { margin-left:0; width:100%; }
            .mobile-menu {
                display:inline-grid; place-items:center; width:40px; height:40px; border:0; border-radius:11px;
                background:#f0edff; color:var(--primary); cursor:pointer; font-size:20px;
            }
            .menu-overlay { position:fixed; inset:0; background:rgba(16,18,37,.38); backdrop-filter:blur(2px); z-index:80; }
            .menu-overlay.show { display:block; }
            .grid { grid-template-columns:1fr; }
            .form-grid { grid-template-columns:1fr; }
            .full { grid-column:auto; }
            .section-header { flex-wrap:wrap; }
        }

        @media (max-width: 600px) {
            .topbar { min-height:66px; padding:10px 13px; gap:10px; }
            .topbar h1 { font-size:17px; line-height:1.25; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; max-width:52vw; }
            .admin-pill { padding:6px; }
            .admin-pill .admin-name { display:none; }
            .avatar { width:31px; height:31px; }
            .logout { padding:8px; }
            .logout span { display:none; }
            .content { padding:13px; }
            .card { padding:16px; border-radius:15px; margin-bottom:14px; }
            .card h2 { font-size:19px; }
            .section-header { align-items:stretch; flex-direction:column; gap:11px; }
            .section-header .btn { width:100%; }
            .search-row { display:grid; grid-template-columns:1fr 1fr; gap:9px; }
            .search-row input { grid-column:1/-1; max-width:none; min-width:0; }
            .search-row .btn { width:100%; }
            input,select,textarea { font-size:16px; }

            .customer-table { min-width:0; display:block; background:transparent; border:0; }
            .customer-table thead { display:none; }
            .customer-table tbody { display:block; }
            .customer-table tr { display:block; border:1px solid var(--border); border-radius:15px; margin-bottom:12px; overflow:hidden; background:white; box-shadow:0 8px 22px rgba(32,34,53,.05); }
            .customer-table td { display:grid; grid-template-columns:115px 1fr; gap:12px; align-items:center; padding:11px 12px; border-bottom:1px solid var(--border); font-size:14px; }
            .customer-table td::before { content:attr(data-label); color:var(--muted); font-size:10px; font-weight:900; text-transform:uppercase; letter-spacing:.4px; }
            .customer-table td:last-child { border-bottom:0; }
            .customer-table .action-buttons { width:100%; flex-direction:column; align-items:stretch; }
            .customer-table .action-buttons .btn,
            .customer-table .action-buttons form,
            .customer-table .action-buttons form .btn { width:100%; }

            body { padding-bottom: env(safe-area-inset-bottom); }
            .content { padding-bottom:96px; }
            .network-banner { bottom:86px; max-width:calc(100vw - 28px); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
            .install-app-btn { padding:8px; }
            .install-app-btn span { display:none; }

            .bottom-nav {
                position:fixed; left:10px; right:10px; bottom:max(10px, env(safe-area-inset-bottom));
                z-index:120; display:grid; grid-template-columns:repeat(6,minmax(0,1fr));
                background:rgba(17,18,37,.95); border:1px solid rgba(255,255,255,.10);
                backdrop-filter:blur(18px); -webkit-backdrop-filter:blur(18px);
                border-radius:19px; padding:6px; box-shadow:0 16px 45px rgba(17,18,37,.28);
            }
            .bottom-nav a {
                min-width:0; display:flex; flex-direction:column; align-items:center; justify-content:center;
                gap:3px; color:#aeb2c8; padding:7px 2px; border-radius:13px; font-size:9px; font-weight:800; white-space:nowrap; overflow:hidden;
            }
            .bottom-nav a.active { color:white; background:linear-gradient(135deg,var(--primary),#7a55ff); }
            .bottom-nav svg { width:19px; height:19px; stroke:currentColor; flex:0 0 auto; }
            .bottom-nav span { display:block; width:100%; text-align:center; overflow:hidden; text-overflow:ellipsis; }

            @media (max-width: 380px) {
                .bottom-nav { left:6px; right:6px; padding:5px; gap:1px; }
                .bottom-nav a { padding:6px 1px; font-size:8px; border-radius:11px; }
                .bottom-nav svg { width:18px; height:18px; }
            }
        }


        /* ---------------- ALL DATA / DATABASE VIEWER ---------------- */
        .data-hero {
            display:grid; grid-template-columns:1.4fr .6fr; gap:18px; align-items:stretch;
            margin-bottom:20px;
        }
        .data-hero-main {
            padding:24px; border-radius:20px; color:white;
            background:linear-gradient(135deg,#17182d,#5b3df5 115%);
            box-shadow:0 18px 45px rgba(43,38,93,.18);
        }
        .data-hero-main h2 { margin:0 0 8px; font-size:26px; }
        .data-hero-main p { margin:0; color:rgba(255,255,255,.76); line-height:1.55; }
        .data-db-card {
            border-radius:20px; padding:22px; background:white; border:1px solid var(--border);
            display:flex; flex-direction:column; justify-content:center; box-shadow:0 10px 30px rgba(28,32,57,.06);
        }
        .data-db-card small { color:var(--muted); font-weight:700; }
        .data-db-card strong { font-size:18px; margin-top:5px; word-break:break-word; }
        .data-summary-grid {
            display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:12px; margin-bottom:20px;
        }
        .data-summary {
            background:white; border:1px solid var(--border); border-radius:16px; padding:16px;
            box-shadow:0 7px 22px rgba(28,32,57,.04);
        }
        .data-summary .num { font-size:25px; font-weight:900; line-height:1; margin-bottom:7px; }
        .data-summary .label { color:var(--muted); font-size:12px; font-weight:800; text-transform:uppercase; }
        .data-section { margin-bottom:20px; }
        .data-section-head {
            display:flex; align-items:center; justify-content:space-between; gap:12px; margin-bottom:14px;
        }
        .data-section-head h2 { margin:0; }
        .data-count {
            display:inline-flex; align-items:center; justify-content:center; min-width:30px; height:30px;
            padding:0 9px; border-radius:999px; background:#eceafb; color:var(--primary-dark); font-size:12px; font-weight:900;
        }
        .mobile-data-cards { display:none; }
        .mobile-data-card {
            background:white; border:1px solid var(--border); border-radius:16px; padding:15px;
            margin-bottom:11px; box-shadow:0 7px 22px rgba(28,32,57,.04);
        }
        .mobile-data-card .row {
            display:grid; grid-template-columns:115px 1fr; gap:10px; padding:6px 0;
            border-bottom:1px dashed #eceef5; align-items:start;
        }
        .mobile-data-card .row:last-child { border-bottom:0; }
        .mobile-data-card .k { color:var(--muted); font-size:11px; font-weight:900; text-transform:uppercase; }
        .mobile-data-card .v { font-size:13px; font-weight:700; overflow-wrap:anywhere; }
        .data-search {
            display:flex; align-items:center; gap:10px; margin-bottom:18px;
        }
        .data-search input { max-width:520px; }
        .privacy-note {
            padding:12px 14px; border-radius:12px; background:#fff8e7; color:#7c5300;
            border:1px solid #f5df9b; font-size:13px; font-weight:700; margin-bottom:18px;
        }

        @media (max-width: 1150px) {
            .data-summary-grid { grid-template-columns:repeat(3,minmax(0,1fr)); }
        }
        @media (max-width: 700px) {
            .data-hero { grid-template-columns:1fr; }
            .data-hero-main h2 { font-size:22px; }
            .data-summary-grid { grid-template-columns:repeat(2,minmax(0,1fr)); }
            .desktop-data-table { display:none; }
            .mobile-data-cards { display:block; }
            .data-search { display:block; }
            .data-search input { max-width:none; margin-bottom:8px; }
            .mobile-data-card .row { grid-template-columns:100px 1fr; }
        }

        @media (max-width: 380px) {
            .customer-table td { grid-template-columns:1fr; gap:5px; }
            .topbar h1 { max-width:44vw; }
        }


        /* ---------------- EXCEL-STYLE DATA STORE ---------------- */
        .excel-toolbar { display:flex; gap:10px; flex-wrap:wrap; align-items:center; margin-bottom:18px; }
        .excel-toolbar .btn svg, .excel-actions .icon-btn svg { width:16px; height:16px; stroke:currentColor; }
        .excel-sheet { background:#fff; border:1px solid #cfd4dc; border-radius:10px; overflow:hidden; margin-bottom:22px; box-shadow:0 8px 20px rgba(31,35,64,.05); }
        .excel-sheet-head { display:flex; align-items:center; justify-content:space-between; gap:12px; padding:10px 12px; background:#f3f5f7; border-bottom:1px solid #cfd4dc; }
        .excel-sheet-title { display:flex; align-items:center; gap:9px; min-width:0; }
        .excel-sheet-title h2 { margin:0; font-size:17px; }
        .excel-grid-wrap { overflow:auto; max-width:100%; }
        .excel-table { width:100%; min-width:900px; border-collapse:collapse; border-spacing:0; background:#fff; }
        .excel-table th, .excel-table td { border:1px solid #d9dde3; padding:7px 9px; font-size:13px; line-height:1.25; white-space:nowrap; text-align:left; }
        .excel-table th { position:sticky; top:0; z-index:2; background:#e9ecef; color:#222; font-size:12px; font-weight:800; text-transform:none; letter-spacing:0; }
        .excel-table tbody tr:nth-child(even) td { background:#fbfcfd; }
        .excel-table tbody tr:hover td { background:#f1f7ff; }
        .excel-table .row-no { width:52px; text-align:center; background:#f3f5f7 !important; color:#606770; font-weight:800; }
        .excel-actions { display:flex; gap:6px; align-items:center; }
        .icon-btn { width:32px; height:32px; display:inline-grid; place-items:center; border-radius:7px; border:1px solid #d8dde5; background:#fff; color:#50586b; cursor:pointer; transition:.15s ease; padding:0; }
        .icon-btn:hover { transform:translateY(-1px); background:#f6f7fb; }
        .icon-btn.view { color:var(--primary-dark); background:#f3f0ff; border-color:#ded7ff; }
        .icon-btn.edit { color:#9a6500; background:#fff8e7; border-color:#f3dfaa; }
        .icon-btn.delete { color:var(--danger); background:#fff0f0; border-color:#ffd6d6; }
        .icon-btn.download { color:var(--success); background:#eaf9f1; border-color:#cceedd; }
        .excel-btn { background:#eaf7ef; color:#167c50; border:1px solid #cce8d8; }
        .excel-btn:hover { background:#dff2e7; }
        @media (max-width:700px) {
            .desktop-data-table { display:block !important; }
            .mobile-data-cards { display:none !important; }
            .excel-sheet { border-radius:8px; }
            .excel-sheet-head { align-items:flex-start; }
            .excel-table { min-width:850px; }
            .excel-table th, .excel-table td { padding:7px 8px; font-size:12px; }
        }

        /* ------------------------- LIVE LOGS ------------------------- */
        .logs-toolbar { display:flex; gap:10px; flex-wrap:wrap; align-items:center; margin-bottom:18px; }
        .logs-live-dot { width:9px; height:9px; border-radius:50%; background:#15945f; display:inline-block; box-shadow:0 0 0 5px rgba(21,148,95,.12); }
        .device-grid { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:14px; margin-bottom:20px; }
        .device-card { background:white; border:1px solid var(--border); border-radius:16px; padding:17px; box-shadow:0 8px 24px rgba(28,32,57,.05); }
        .device-card-head { display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:10px; }
        .device-name { font-size:16px; font-weight:900; }
        .device-meta { color:var(--muted); font-size:12px; line-height:1.65; overflow-wrap:anywhere; }
        .status-online { color:#137a4f; background:#e9f9f1; border:1px solid #cdeedd; }
        .status-offline { color:#697087; background:#f2f4f8; border:1px solid #e2e5ec; }
        .log-status { display:inline-flex; align-items:center; gap:6px; border-radius:999px; padding:5px 8px; font-size:10px; font-weight:900; }
        .log-table { min-width:1050px; }
        .log-table td { vertical-align:middle; }
        .log-action { font-weight:800; }
        .log-ip { font-family:ui-monospace,SFMono-Regular,Menlo,monospace; font-size:12px; }
        @media (max-width:900px) { .device-grid { grid-template-columns:1fr 1fr; } }
        @media (max-width:600px) {
            .device-grid { grid-template-columns:1fr; }
            .logs-toolbar .btn { flex:1 1 auto; }
            .log-table { min-width:980px; }
        }

        /* Instant feedback for every internal link/form action. */
        .fast-progress {
            position:fixed; top:0; left:0; width:0; height:3px; z-index:9999;
            background:linear-gradient(90deg,var(--primary),#9d83ff);
            opacity:0; transition:width .16s ease, opacity .18s ease;
            pointer-events:none;
        }
        .fast-progress.show { opacity:1; width:72%; }
        .fast-progress.done { width:100%; opacity:0; }
        .content.fast-loading { opacity:.72; transition:opacity .12s ease; pointer-events:none; }
        button.fast-busy, .btn.fast-busy { opacity:.72; pointer-events:none; }
    </style>
</head>
<body>
{% if logged_in %}
<div class="layout">
    <aside class="sidebar" id="sidebar">
        <div class="brand">
            <div class="brand-logo">A</div>
            <div>AMBAAL <span>SHOP</span></div>
        </div>

        <div class="nav-label">Main</div>
        <a class="nav-link {{ 'active' if active_page == 'dashboard' else '' }}" href="{{ url_for('dashboard') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 13h8V3H3v10Zm10 8h8V11h-8v10ZM3 21h8v-6H3v6Zm10-12h8V3h-8v6Z"/></svg>
            <span>Dashboard</span>
        </a>

        <div class="nav-label">Management</div>
        <a class="nav-link {{ 'active' if active_page == 'loans' else '' }}" href="{{ url_for('customer_loans') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M19 8v6M22 11h-6"/></svg>
            <span>Customer Loans</span>
        </a>
        <a class="nav-link {{ 'active' if active_page == 'items' else '' }}" href="{{ url_for('shop_items') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="m21 16-4 4-4-4"/><path d="M17 20V4"/><path d="M3 8h10M3 12h8M3 16h6"/></svg>
            <span>Shop Things</span>
        </a>
        <a class="nav-link {{ 'active' if active_page == 'prices' else '' }}" href="{{ url_for('price_management') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M20 13V7a2 2 0 0 0-2-2H6a2 2 0 0 0-2 2v10a2 2 0 0 0 2 2h7"/><path d="M16 16h6M19 13v6"/><circle cx="9" cy="10" r="2"/></svg>
            <span>Prices</span>
        </a>
        <a class="nav-link {{ 'active' if active_page == 'database' else '' }}" href="{{ url_for('all_data') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><ellipse cx="12" cy="5" rx="8" ry="3"/><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"/><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"/></svg>
            <span>All Data</span>
        </a>
        <a class="nav-link {{ 'active' if active_page == 'logs' else '' }}" href="{{ url_for('live_logs') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 3v18h18"/><path d="m7 15 4-4 3 3 5-7"/><circle cx="7" cy="15" r="1"/><circle cx="11" cy="11" r="1"/><circle cx="14" cy="14" r="1"/><circle cx="19" cy="7" r="1"/></svg>
            <span>Live Logs</span>
        </a>

        <div class="sidebar-footer">Simple, fast and responsive shop management for desktop, tablet and mobile.</div>
    </aside>
    <div class="menu-overlay" id="menuOverlay" onclick="closeMenu()"></div>

    <main class="main">
        <header class="topbar">
            <div class="topbar-left">
                <button class="mobile-menu" onclick="toggleMenu()">☰</button>
                <h1>{{ page_heading }}</h1>
            </div>
            <div class="user-area">
                <button class="install-app-btn" id="installAppBtn" type="button" title="Install Ambaal Shop app">
                    <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg>
                    <span>Install App</span>
                </button>
                <div class="admin-pill">
                    <div class="avatar">{{ (session.get('username') or 'A')[0]|upper }}</div>
                    <span class="admin-name">{{ session.get('username') }}</span>
                </div>
                <a class="logout" href="{{ url_for('logout') }}">
                    <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><path d="m16 17 5-5-5-5"/><path d="M21 12H9"/></svg>
                    <span>Logout</span>
                </a>
            </div>
        </header>

        <section class="content">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }}">{{ message }}</div>
                {% endfor %}
            {% endwith %}

            {{ content|safe }}
        </section>
    </main>

    <nav class="bottom-nav" aria-label="Mobile navigation">
        <a class="{{ 'active' if active_page == 'dashboard' else '' }}" href="{{ url_for('dashboard') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 13h8V3H3v10Zm10 8h8V11h-8v10ZM3 21h8v-6H3v6Zm10-12h8V3h-8v6Z"/></svg><span>Home</span>
        </a>
        <a class="{{ 'active' if active_page == 'loans' else '' }}" href="{{ url_for('customer_loans') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M19 8v6M22 11h-6"/></svg><span>Loans</span>
        </a>
        <a class="{{ 'active' if active_page == 'items' else '' }}" href="{{ url_for('shop_items') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M4 6h16M4 12h16M4 18h10"/></svg><span>Items</span>
        </a>
        <a class="{{ 'active' if active_page == 'prices' else '' }}" href="{{ url_for('price_management') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M20 13V7a2 2 0 0 0-2-2H6a2 2 0 0 0-2 2v10a2 2 0 0 0 2 2h7"/><path d="M16 16h6M19 13v6"/></svg><span>Prices</span>
        </a>
        <a class="{{ 'active' if active_page == 'database' else '' }}" href="{{ url_for('all_data') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><ellipse cx="12" cy="5" rx="8" ry="3"/><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"/><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"/></svg><span>Data</span>
        </a>
        <a class="{{ 'active' if active_page == 'logs' else '' }}" href="{{ url_for('live_logs') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 3v18h18"/><path d="m7 15 4-4 3 3 5-7"/></svg><span>Logs</span>
        </a>
    </nav>
</div>
{% else %}
<div class="login-page">
    <div class="login-card">
        <h1>Ambaal Shop</h1>
        <p>Sign in to open the administration system.</p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% for category, message in messages %}
                <div class="alert alert-{{ category }}">{{ message }}</div>
            {% endfor %}
        {% endwith %}

        {{ content|safe }}
    </div>
</div>
{% endif %}

<div class="fast-progress" id="fastProgress" aria-hidden="true"></div>
<div class="network-banner" id="networkBanner" role="status" aria-live="polite"></div>

<script>
    function toggleMenu() {
        const sidebar = document.getElementById("sidebar");
        const overlay = document.getElementById("menuOverlay");
        sidebar.classList.toggle("open");
        if (overlay) overlay.classList.toggle("show");
    }

    function closeMenu() {
        const sidebar = document.getElementById("sidebar");
        const overlay = document.getElementById("menuOverlay");
        sidebar.classList.remove("open");
        if (overlay) overlay.classList.remove("show");
    }

    function confirmDelete(message) {
        return confirm(message || "Are you sure?");
    }

    // PWA installation prompt (supported Chromium browsers).
    let deferredInstallPrompt = null;
    const installButton = document.getElementById("installAppBtn");

    window.addEventListener("beforeinstallprompt", (event) => {
        event.preventDefault();
        deferredInstallPrompt = event;
        if (installButton) installButton.classList.add("show");
    });

    if (installButton) {
        installButton.addEventListener("click", async () => {
            if (!deferredInstallPrompt) return;
            deferredInstallPrompt.prompt();
            await deferredInstallPrompt.userChoice;
            deferredInstallPrompt = null;
            installButton.classList.remove("show");
        });
    }

    window.addEventListener("appinstalled", () => {
        deferredInstallPrompt = null;
        if (installButton) installButton.classList.remove("show");
    });

    // -----------------------------------------------------------------\n    // UNIVERSAL FAST RESPONSE ENGINE\n    // Applies automatically to every safe same-origin link and normal form.\n    // Pages are kept in short-lived memory, likely destinations are prefetched,\n    // and only the changing page shell is swapped instead of a browser reload.\n    // -----------------------------------------------------------------\n    const fastPageCache = new Map();\n    const FAST_CACHE_MS = 20000;\n    let fastNavigationController = null;\n\n    function fastProgressStart(trigger) {\n        const bar = document.getElementById("fastProgress");\n        const content = document.querySelector(".content");\n        if (bar) { bar.className = "fast-progress show"; }\n        if (content) content.classList.add("fast-loading");\n        if (trigger) trigger.classList.add("fast-busy");\n    }\n\n    function fastProgressDone() {\n        const bar = document.getElementById("fastProgress");\n        const content = document.querySelector(".content");\n        if (content) content.classList.remove("fast-loading");\n        document.querySelectorAll(".fast-busy").forEach(el => el.classList.remove("fast-busy"));\n        if (bar) {\n            bar.className = "fast-progress done";\n            setTimeout(() => { bar.className = "fast-progress"; bar.style.width = ""; }, 220);\n        }\n    }\n\n    function fastCacheSet(url, html) {\n        fastPageCache.set(url, { html, time: Date.now() });\n        if (fastPageCache.size > 18) {\n            const first = fastPageCache.keys().next().value;\n            fastPageCache.delete(first);\n        }\n    }\n\n    function fastCacheGet(url) {\n        const item = fastPageCache.get(url);\n        if (!item) return null;\n        if (Date.now() - item.time > FAST_CACHE_MS) {\n            fastPageCache.delete(url);\n            return null;\n        }\n        return item.html;\n    }\n\n    function fastCanHandleUrl(rawUrl) {\n        try {\n            const u = new URL(rawUrl, location.href);\n            if (u.origin !== location.origin) return false;\n            if (u.pathname === "/logout" || u.pathname.startsWith("/all-data/export-excel")) return false;\n            if (/\\.(?:xlsx?|csv|pdf|zip|png|jpe?g|webp)$/i.test(u.pathname)) return false;\n            return true;\n        } catch (_) { return false; }\n    }\n\n    function fastApplyDocument(html, url, pushHistory=true) {\n        const parsed = new DOMParser().parseFromString(html, "text/html");\n        const incomingMain = parsed.querySelector(".main");\n        const currentMain = document.querySelector(".main");\n        const incomingSidebar = parsed.querySelector(".sidebar");\n        const currentSidebar = document.querySelector(".sidebar");\n        const incomingBottom = parsed.querySelector(".bottom-nav");\n        const currentBottom = document.querySelector(".bottom-nav");\n\n        if (!incomingMain || !currentMain) {\n            location.href = url;\n            return;\n        }\n\n        currentMain.innerHTML = incomingMain.innerHTML;\n        if (incomingSidebar && currentSidebar) currentSidebar.innerHTML = incomingSidebar.innerHTML;\n        if (incomingBottom && currentBottom) currentBottom.innerHTML = incomingBottom.innerHTML;\n        document.title = parsed.title || document.title;\n        closeMenu();\n        if (pushHistory) history.pushState({ fast: true }, "", url);\n        window.scrollTo({ top: 0, behavior: "instant" });\n        fastProgressDone();\n    }\n\n    async function fastFetchPage(url, options={}, trigger=null, pushHistory=true) {\n        const absolute = new URL(url, location.href).href;\n        const method = (options.method || "GET").toUpperCase();\n\n        // GET pages that were recently visited/prefetched appear immediately.\n        if (method === "GET") {\n            const cached = fastCacheGet(absolute);\n            if (cached) {\n                fastProgressStart(trigger);\n                fastApplyDocument(cached, absolute, pushHistory);\n                // Refresh quietly so future visits stay current.\n                fetch(absolute, { headers: { "X-Fast-Navigation": "1" }, credentials: "same-origin" })\n                    .then(r => r.ok ? r.text() : Promise.reject())\n                    .then(text => fastCacheSet(rURL(text, absolute), text))\n                    .catch(() => {});\n                return;\n            }\n        }\n\n        fastProgressStart(trigger);\n        if (fastNavigationController) fastNavigationController.abort();\n        fastNavigationController = new AbortController();\n\n        const headers = new Headers(options.headers || {});\n        headers.set("X-Fast-Navigation", "1");\n        try {\n            const response = await fetch(absolute, {\n                ...options, headers, credentials: "same-origin",\n                signal: fastNavigationController.signal\n            });\n            const html = await response.text();\n            const finalUrl = response.url || absolute;\n            if (response.ok && response.headers.get("content-type")?.includes("text/html")) {\n                if (method === "GET") fastCacheSet(finalUrl, html);\n                else fastPageCache.clear(); // a write may change any displayed data\n                fastApplyDocument(html, finalUrl, pushHistory);\n            } else {\n                location.href = finalUrl;\n            }\n        } catch (error) {\n            if (error.name !== "AbortError") location.href = absolute;\n        } finally {\n            fastNavigationController = null;\n            fastProgressDone();\n        }\n    }\n\n    // Helper used only by the quiet background refresh above.\n    function rURL(_html, fallback) { return fallback; }\n\n    // Every safe same-origin anchor becomes fast automatically.\n    document.addEventListener("click", (event) => {\n        const link = event.target.closest("a[href]");\n        if (!link) return;\n        if (event.defaultPrevented || event.button !== 0 || event.metaKey || event.ctrlKey || event.shiftKey || event.altKey) return;\n        if (link.target === "_blank" || link.hasAttribute("download")) return;\n        if (!fastCanHandleUrl(link.href)) return;\n        event.preventDefault();\n        fastFetchPage(link.href, { method: "GET" }, link, true);\n    });\n\n    // Search, Add, Update, Delete, Finish, Reopen and other forms all use the\n    // same fast engine. Browser validation and existing confirm dialogs remain.\n    document.addEventListener("submit", (event) => {\n        const form = event.target;\n        if (!(form instanceof HTMLFormElement)) return;\n        if (form.target === "_blank" || form.dataset.noFast === "1") return;\n        const action = form.action || location.href;\n        if (!fastCanHandleUrl(action)) return;\n        if (!form.reportValidity()) { event.preventDefault(); return; }\n        event.preventDefault();\n\n        const method = (form.method || "GET").toUpperCase();\n        const submitter = event.submitter || form.querySelector('[type="submit"]');\n        if (method === "GET") {\n            const u = new URL(action, location.href);\n            new FormData(form).forEach((value, key) => { if (String(value).length) u.searchParams.set(key, value); });\n            fastFetchPage(u.href, { method: "GET" }, submitter, true);\n        } else {\n            fastFetchPage(action, { method, body: new FormData(form) }, submitter, true);\n        }\n    });\n\n    // Browser Back/Forward also swaps through the fast engine.\n    window.addEventListener("popstate", () => {\n        if (fastCanHandleUrl(location.href)) fastFetchPage(location.href, { method: "GET" }, null, false);\n    });\n\n    // Prefetch the six main destinations after the current screen is usable.\n    // This happens quietly and makes bottom/sidebar navigation much faster.\n    function fastPrefetchMainPages() {\n        const links = [...document.querySelectorAll('.bottom-nav a[href], .sidebar a.nav-link[href]')];\n        const unique = [...new Set(links.map(a => a.href).filter(fastCanHandleUrl))];\n        unique.forEach((url, index) => {\n            setTimeout(() => {\n                if (fastCacheGet(url)) return;\n                fetch(url, { headers: { "X-Fast-Prefetch": "1" }, credentials: "same-origin" })\n                    .then(r => (r.ok && r.headers.get("content-type")?.includes("text/html")) ? r.text() : Promise.reject())\n                    .then(html => fastCacheSet(url, html))\n                    .catch(() => {});\n            }, 250 + index * 120);\n        });\n    }\n    if (document.querySelector(".main")) {\n        if ("requestIdleCallback" in window) requestIdleCallback(fastPrefetchMainPages, { timeout: 1200 });\n        else setTimeout(fastPrefetchMainPages, 500);\n    }\n\n    // Network status feedback. Database writes still require the Flask server.
    const networkBanner = document.getElementById("networkBanner");
    let networkTimer = null;
    function showNetworkState(isOnline) {
        if (!networkBanner) return;
        clearTimeout(networkTimer);
        networkBanner.className = "network-banner " + (isOnline ? "online" : "offline");
        networkBanner.textContent = isOnline
            ? "Back online"
            : "Offline — viewing cached app shell only";
        if (isOnline) {
            networkTimer = setTimeout(() => {
                networkBanner.className = "network-banner";
            }, 2500);
        }
    }
    window.addEventListener("online", () => showNetworkState(true));
    window.addEventListener("offline", () => showNetworkState(false));
    if (!navigator.onLine) showNetworkState(false);

    // Register service worker for installability and offline app shell.
    if ("serviceWorker" in navigator) {
        window.addEventListener("load", () => {
            navigator.serviceWorker.register("{{ url_for('service_worker') }}")
                .catch((error) => console.log("Service worker registration failed:", error));
        });
    }
</script>
</body>
</html>
"""


@lru_cache(maxsize=64)
def _compiled_template(source):
    """Compile each static template only once per worker process."""
    return app.jinja_env.from_string(source)


def _render_cached(source, **context):
    # Match Flask's render_template_string behaviour by including normal
    # request/session context processors, while avoiding repeated compilation.
    app.update_template_context(context)
    return _compiled_template(source).render(context)


def render_page(title, page_heading, content_template, active_page="", **context):
    content = _render_cached(content_template, **context)
    return _render_cached(
        BASE_TEMPLATE,
        title=title,
        page_heading=page_heading,
        content=content,
        active_page=active_page,
        logged_in=("user_id" in session)
    )


# ---------------------------- PWA -----------------------------

@app.route("/manifest.webmanifest")
def pwa_manifest():
    """Web-app manifest used when Ambaal Shop is installed as a PWA."""
    return jsonify({
        "name": "Ambaal Shop Management",
        "short_name": "Ambaal Shop",
        "description": "Customer loans, shop items and price management.",
        "start_url": "/dashboard",
        "scope": "/",
        "display": "standalone",
        "background_color": "#111225",
        "theme_color": "#111225",
        "orientation": "any",
        "icons": [
            {
                "src": url_for("pwa_icon", size=192),
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": url_for("pwa_icon", size=512),
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    })


@app.route("/pwa-icon-<int:size>.png")
def pwa_icon(size):
    """Serve embedded app icons without needing a static folder."""
    import base64
    if size == 192:
        raw = base64.b64decode(PWA_ICON_192)
    elif size == 512:
        raw = base64.b64decode(PWA_ICON_512)
    else:
        return Response(status=404)
    response = Response(raw, mimetype="image/png")
    response.headers["Cache-Control"] = "public, max-age=604800"
    return response


@app.route("/offline")
def offline_page():
    """Small cached page shown when the hosted server cannot be reached."""
    html = r"""
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
        <meta name="theme-color" content="#111225">
        <title>Ambaal Shop - Offline</title>
        <style>
            *{box-sizing:border-box}body{margin:0;min-height:100vh;display:grid;place-items:center;padding:24px;
            font-family:Inter,system-ui,-apple-system,"Segoe UI",sans-serif;background:linear-gradient(145deg,#111225,#26284d);color:#fff}
            .box{width:min(460px,100%);padding:32px;border-radius:26px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12);
            box-shadow:0 24px 70px rgba(0,0,0,.25);text-align:center}.logo{width:72px;height:72px;margin:0 auto 18px;border-radius:22px;display:grid;
            place-items:center;background:linear-gradient(135deg,#6d4aff,#9f83ff);font-size:36px;font-weight:900}.muted{color:#c5c8db;line-height:1.6}
            button{border:0;border-radius:13px;padding:13px 18px;background:#6d4aff;color:white;font-weight:800;font-size:15px;cursor:pointer}
        </style>
    </head>
    <body><div class="box"><div class="logo">A</div><h1>You're offline</h1>
    <p class="muted">Ambaal Shop is installed, but customer records and database actions need a connection to the Flask server.</p>
    <button onclick="location.reload()">Try Again</button></div></body></html>
    """
    return Response(html, mimetype="text/html")


@app.route("/service-worker.js")
def service_worker():
    """Conservative service worker: cache only the app shell assets, not customer data."""
    script = r"""
    const CACHE = 'ambaal-shop-shell-v1';
    const SHELL = ['/offline', '/manifest.webmanifest', '/pwa-icon-192.png', '/pwa-icon-512.png'];

    self.addEventListener('install', event => {
      event.waitUntil(caches.open(CACHE).then(cache => cache.addAll(SHELL)).then(() => self.skipWaiting()));
    });

    self.addEventListener('activate', event => {
      event.waitUntil(
        caches.keys().then(keys => Promise.all(keys.filter(k => k !== CACHE).map(k => caches.delete(k))))
          .then(() => self.clients.claim())
      );
    });

    self.addEventListener('fetch', event => {
      if (event.request.method !== 'GET') return;
      const url = new URL(event.request.url);
      if (url.origin !== self.location.origin) return;

      // Never cache authenticated application pages or database responses.
      if (event.request.mode === 'navigate') {
        event.respondWith(fetch(event.request).catch(() => caches.match('/offline')));
        return;
      }

      if (SHELL.includes(url.pathname)) {
        event.respondWith(caches.match(event.request).then(cached => cached || fetch(event.request)));
      }
    });
    """
    response = Response(script, mimetype="application/javascript")
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


# --------------------------- LOGIN ----------------------------

@app.route("/", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        connection = None
        cursor = None
        try:
            connection = db_connection()
            cursor = connection.cursor(dictionary=True)
            cursor.execute(
                "SELECT id, username, password_hash FROM users WHERE username = %s",
                (username,)
            )
            user = cursor.fetchone()

            if user and check_password_hash(user["password_hash"], password):
                session.clear()
                session["user_id"] = user["id"]
                session["username"] = user["username"]
                log_activity("Login successful", user_id=user["id"], username=user["username"])
                flash("Login successful.", "success")
                return redirect(url_for("dashboard"))

            flash("Invalid username or password.", "danger")

        except Error as error:
            flash(f"Database error: {error}", "danger")
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()

    login_form = r"""
    <form method="POST">
        <div class="form-group">
            <label>Username</label>
            <input type="text" name="username" required autocomplete="username"
                   placeholder="Enter username">
        </div>

        <div class="form-group">
            <label>Password</label>
            <input type="password" name="password" required
                   autocomplete="current-password" placeholder="Enter password">
        </div>

        <button class="btn btn-primary" type="submit">Login</button>
    </form>
    """
    return render_page("Login", "Login", login_form)


@app.route("/logout")
def logout():
    if "user_id" in session:
        log_activity("Logged out")
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ------------------------- DASHBOARD --------------------------

@app.route("/dashboard")
@login_required
def dashboard():
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)

        # Fetch all dashboard counters in one round-trip to the hosted DB.
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM customers) AS customer_count,
                (SELECT COUNT(*) FROM shop_items) AS item_count,
                COALESCE((
                    SELECT SUM(CASE
                        WHEN transaction_type = 'LOAN' THEN amount
                        WHEN transaction_type = 'PAYMENT' THEN -amount
                        ELSE 0
                    END)
                    FROM loan_transactions
                ), 0) AS total_balance
        """)
        dashboard_stats = cursor.fetchone()
        customer_count = dashboard_stats["customer_count"]
        item_count = dashboard_stats["item_count"]
        total_balance = dashboard_stats["total_balance"]

        cursor.execute("""
            SELECT
                lt.id,
                c.customer_name,
                c.customer_code,
                lt.transaction_date,
                lt.transaction_type,
                lt.amount
            FROM loan_transactions lt
            JOIN customers c ON c.id = lt.customer_id
            ORDER BY lt.id DESC
            LIMIT 8
        """)
        recent_transactions = cursor.fetchall()

    except Error as error:
        flash(f"Database error: {error}", "danger")
        customer_count = 0
        item_count = 0
        total_balance = 0
        recent_transactions = []
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="grid">
        <div class="card">
            <div class="stat-title">Registered Customers</div>
            <div class="stat-value">{{ customer_count }}</div>
        </div>
        <div class="card">
            <div class="stat-title">Total Outstanding Loan</div>
            <div class="stat-value">Rs. {{ total_balance|money }}</div>
        </div>
        <div class="card">
            <div class="stat-title">Shop Items</div>
            <div class="stat-value">{{ item_count }}</div>
        </div>
    </div>

    <div class="card">
        <div class="section-header">
            <h2>Recent Loan Activity</h2>
            <a class="btn btn-secondary" href="{{ url_for('customer_loans') }}">
                Open Loan Management
            </a>
        </div>

        {% if recent_transactions %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Customer</th>
                        <th>Customer ID</th>
                        <th>Type</th>
                        <th>Amount</th>
                    </tr>
                </thead>
                <tbody>
                {% for transaction in recent_transactions %}
                    <tr>
                        <td>{{ transaction.transaction_date }}</td>
                        <td>{{ transaction.customer_name }}</td>
                        <td>{{ transaction.customer_code }}</td>
                        <td>
                            <span class="badge {{ 'badge-loan' if transaction.transaction_type == 'LOAN' else 'badge-payment' }}">
                                {{ transaction.transaction_type }}
                            </span>
                        </td>
                        <td>Rs. {{ transaction.amount|money }}</td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No loan transactions have been recorded.</div>
        {% endif %}
    </div>
    """

    return render_page(
        "Dashboard",
        "Dashboard",
        template,
        active_page="dashboard",
        customer_count=customer_count,
        item_count=item_count,
        total_balance=total_balance,
        recent_transactions=recent_transactions
    )


# ---------------------- CUSTOMER LOANS ------------------------

@app.route("/customer-loans")
@login_required
def customer_loans():
    search = request.args.get("search", "").strip()

    connection = None
    cursor = None
    customers = []
    transactions = []

    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)

        customer_query = """
            SELECT
                c.id,
                c.customer_code,
                c.customer_name,
                c.mobile_number,
                c.created_at,
                COALESCE(SUM(
                    CASE
                        WHEN lt.transaction_type = 'LOAN' THEN lt.amount
                        WHEN lt.transaction_type = 'PAYMENT' THEN -lt.amount
                        ELSE 0
                    END
                ), 0) AS balance
            FROM customers c
            LEFT JOIN loan_transactions lt ON lt.customer_id = c.id
        """
        parameters = []

        if search:
            customer_query += """
                WHERE c.customer_name LIKE %s
                   OR c.customer_code LIKE %s
                   OR c.mobile_number LIKE %s
            """
            pattern = f"%{search}%"
            parameters.extend([pattern, pattern, pattern])

        customer_query += """
            GROUP BY c.id, c.customer_code, c.customer_name,
                     c.mobile_number, c.created_at
            ORDER BY c.customer_name
        """
        cursor.execute(customer_query, tuple(parameters))
        customers = cursor.fetchall()

        cursor.execute("""
            SELECT
                lt.id,
                lt.transaction_date,
                lt.transaction_type,
                lt.amount,
                lt.note,
                lt.created_at,
                c.id AS customer_id,
                c.customer_code,
                c.customer_name,
                c.mobile_number
            FROM loan_transactions lt
            JOIN customers c ON c.id = lt.customer_id
            ORDER BY lt.transaction_date DESC, lt.id DESC
            LIMIT 100
        """)
        transactions = cursor.fetchall()

    except Error as error:
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="card">
        <div class="section-header">
            <h2>Customer Accounts</h2>
            <a class="btn btn-primary" href="{{ url_for('add_customer') }}">
                <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                    <path d="M12 5v14M5 12h14"/>
                </svg>
                Add New Customer
            </a>
        </div>

        <form class="search-row" method="GET">
            <input type="text" name="search" value="{{ search }}"
                   placeholder="Search name, customer ID or mobile number">
            <button class="btn btn-secondary" type="submit">
                <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                    <circle cx="11" cy="11" r="7"/>
                    <path d="m20 20-3.5-3.5"/>
                </svg>
                Search
            </button>
            {% if search %}
                <a class="btn btn-danger" href="{{ url_for('customer_loans') }}">
                    <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                        <path d="M18 6 6 18M6 6l12 12"/>
                    </svg>
                    Clear
                </a>
            {% endif %}
        </form>

        {% if customers %}
        <div class="table-wrap">
            <table class="customer-table">
                <thead>
                    <tr>
                        <th>Customer ID</th>
                        <th>Name</th>
                        <th>Mobile</th>
                        <th>Current Balance</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                {% for customer in customers %}
                    <tr>
                        <td data-label="Customer ID">{{ customer.customer_code }}</td>
                        <td data-label="Name">{{ customer.customer_name }}</td>
                        <td data-label="Mobile">{{ customer.mobile_number }}</td>
                        <td data-label="Current Balance"
                            class="{{ 'balance-positive' if customer.balance > 0 else 'balance-zero' }}">
                            Rs. {{ customer.balance|money }}
                        </td>
                        <td data-label="Actions">
                            <div class="action-buttons">
                                <a class="btn btn-success"
                                   href="{{ url_for('customer_details', customer_id=customer.id) }}">
                                    <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                                        <path d="M12 5v14M5 12h14"/>
                                    </svg>
                                    View / Add Amount
                                </a>

                                <a class="btn btn-secondary"
                                   href="{{ url_for('edit_customer', customer_id=customer.id) }}">Edit</a>

                                <form method="POST"
                                      action="{{ url_for('delete_customer', customer_id=customer.id) }}"
                                      onsubmit="return confirmDelete('Delete this customer and all loan history? This cannot be undone.');">
                                    <button class="btn btn-danger" type="submit">
                                        <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                                            <path d="M3 6h18"/>
                                            <path d="M8 6V4h8v2"/>
                                            <path d="M19 6l-1 14H6L5 6"/>
                                            <path d="M10 11v5M14 11v5"/>
                                        </svg>
                                        Delete
                                    </button>
                                </form>
                            </div>
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No customers found.</div>
        {% endif %}
    </div>

    <div class="card">
        <h2>All Transaction History</h2>
        {% if transactions %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Customer</th>
                        <th>Customer ID</th>
                        <th>Type</th>
                        <th>Amount</th>
                        <th>Note</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                {% for transaction in transactions %}
                    <tr>
                        <td>{{ transaction.transaction_date }}</td>
                        <td>{{ transaction.customer_name }}</td>
                        <td>{{ transaction.customer_code }}</td>
                        <td>
                            <span class="badge {{ 'badge-loan' if transaction.transaction_type == 'LOAN' else 'badge-payment' }}">
                                {{ transaction.transaction_type }}
                            </span>
                        </td>
                        <td>Rs. {{ transaction.amount|money }}</td>
                        <td>{{ transaction.note or '-' }}</td>
                        <td>
                            <a class="btn btn-secondary"
                               href="{{ url_for('edit_transaction', transaction_id=transaction.id) }}">Edit</a>
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No transaction history available.</div>
        {% endif %}
    </div>
    """

    return render_page(
        "Customer Loans",
        "Customer Loan Management",
        template,
        active_page="loans",
        customers=customers,
        transactions=transactions,
        search=search
    )


@app.route("/customer-loans/customer/<int:customer_id>/edit", methods=["GET", "POST"])
@login_required
def edit_customer(customer_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM customers WHERE id = %s", (customer_id,))
        customer = cursor.fetchone()

        if not customer:
            flash("Customer not found.", "danger")
            return redirect(url_for("customer_loans"))

        if request.method == "POST":
            customer_name = request.form.get("customer_name", "").strip()
            customer_code = request.form.get("customer_code", "").strip()
            mobile_number = request.form.get("mobile_number", "").strip()

            if not customer_name or not customer_code or not mobile_number:
                flash("Customer name, customer ID and mobile number are required.", "danger")
            else:
                try:
                    cursor.execute("""
                        UPDATE customers
                        SET customer_code = %s, customer_name = %s, mobile_number = %s
                        WHERE id = %s
                    """, (customer_code, customer_name, mobile_number, customer_id))
                    connection.commit()
                    flash("Customer updated successfully.", "success")
                    return redirect(url_for("customer_details", customer_id=customer_id))
                except Error as error:
                    connection.rollback()
                    if error.errno == 1062:
                        flash("That customer ID already exists.", "danger")
                    else:
                        raise

        template = r"""
        <div class="card">
            <div class="section-header">
                <h2>Edit Customer</h2>
                <a class="btn btn-secondary" href="{{ url_for('customer_details', customer_id=customer.id) }}">Cancel</a>
            </div>
            <form method="POST">
                <div class="form-grid">
                    <div class="form-group">
                        <label>Customer Name</label>
                        <input type="text" name="customer_name" value="{{ customer.customer_name }}" required>
                    </div>
                    <div class="form-group">
                        <label>Customer ID</label>
                        <input type="text" name="customer_code" value="{{ customer.customer_code }}" required>
                    </div>
                    <div class="form-group">
                        <label>Mobile Number</label>
                        <input type="tel" name="mobile_number" value="{{ customer.mobile_number }}" required>
                    </div>
                    <div class="full">
                        <button class="btn btn-primary" type="submit">Update Customer</button>
                    </div>
                </div>
            </form>
        </div>
        """
        return render_page("Edit Customer", "Edit Customer", template,
                           active_page="loans", customer=customer)
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
        return redirect(url_for("customer_loans"))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route("/customer-loans/customer/<int:customer_id>/delete", methods=["POST"])
@login_required
def delete_customer(customer_id):
    """Delete a customer. Related loan transactions are removed by ON DELETE CASCADE."""
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute(
            "SELECT customer_name FROM customers WHERE id = %s",
            (customer_id,)
        )
        customer = cursor.fetchone()

        if not customer:
            flash("Customer not found.", "danger")
            return redirect(url_for("customer_loans"))

        cursor.execute("DELETE FROM customers WHERE id = %s", (customer_id,))
        connection.commit()
        flash(f"Customer '{customer['customer_name']}' deleted successfully.", "success")

    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    return redirect(url_for("customer_loans"))


@app.route("/customer-loans/add-customer", methods=["GET", "POST"])
@login_required
def add_customer():
    if request.method == "POST":
        customer_name = request.form.get("customer_name", "").strip()
        customer_code = request.form.get("customer_code", "").strip()
        mobile_number = request.form.get("mobile_number", "").strip()
        transaction_date = request.form.get("transaction_date", "")
        initial_loan = request.form.get("initial_loan", "0").strip()
        note = request.form.get("note", "").strip()

        if not customer_name or not customer_code or not mobile_number:
            flash("Customer name, customer ID and mobile number are required.", "danger")
        else:
            connection = None
            cursor = None
            try:
                amount = Decimal(initial_loan or "0")
                if amount < 0:
                    raise InvalidOperation

                connection = db_connection()
                cursor = connection.cursor()
                cursor.execute("""
                    INSERT INTO customers
                        (customer_code, customer_name, mobile_number)
                    VALUES (%s, %s, %s)
                """, (customer_code, customer_name, mobile_number))

                customer_id = cursor.lastrowid

                if amount > 0:
                    cursor.execute("""
                        INSERT INTO loan_transactions
                            (customer_id, transaction_date,
                             transaction_type, amount, note)
                        VALUES (%s, %s, 'LOAN', %s, %s)
                    """, (
                        customer_id,
                        transaction_date or date.today().isoformat(),
                        amount,
                        note or "Initial loan"
                    ))

                connection.commit()
                flash("Customer and initial loan saved successfully.", "success")
                return redirect(
                    url_for("customer_details", customer_id=customer_id)
                )

            except InvalidOperation:
                flash("Loan amount must be a valid positive number.", "danger")
            except Error as error:
                if connection:
                    connection.rollback()
                if error.errno == 1062:
                    flash("That customer ID already exists.", "danger")
                else:
                    flash(f"Database error: {error}", "danger")
            finally:
                if cursor:
                    cursor.close()
                if connection and connection.is_connected():
                    connection.close()

    template = r"""
    <div class="card">
        <div class="section-header">
            <h2>Add New Customer</h2>
            <a class="btn btn-secondary" href="{{ url_for('customer_loans') }}">Back</a>
        </div>

        <form method="POST">
            <div class="form-grid">
                <div class="form-group">
                    <label>Customer Name</label>
                    <input type="text" name="customer_name" required>
                </div>

                <div class="form-group">
                    <label>Customer ID</label>
                    <input type="text" name="customer_code" required
                           placeholder="Example: CUST-001">
                </div>

                <div class="form-group">
                    <label>Mobile Number</label>
                    <input type="tel" name="mobile_number" required>
                </div>

                <div class="form-group">
                    <label>Date</label>
                    <input type="date" name="transaction_date"
                           value="{{ today }}" required>
                </div>

                <div class="form-group">
                    <label>Initial Loan Amount (Rs.)</label>
                    <input type="number" name="initial_loan" min="0"
                           step="0.01" value="0" required>
                </div>

                <div class="form-group">
                    <label>Note</label>
                    <input type="text" name="note"
                           placeholder="Items bought or other details">
                </div>

                <div class="full">
                    <button class="btn btn-primary" type="submit">
                        Save Customer
                    </button>
                </div>
            </div>
        </form>
    </div>
    """

    return render_page(
        "Add Customer",
        "Add New Customer",
        template,
        active_page="loans",
        today=date.today().isoformat()
    )


@app.route("/customer-loans/customer/<int:customer_id>")
@login_required
def customer_details(customer_id):
    connection = None
    cursor = None

    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM customers WHERE id = %s",
            (customer_id,)
        )
        customer = cursor.fetchone()

        if not customer:
            flash("Customer not found.", "danger")
            return redirect(url_for("customer_loans"))

        balance = get_customer_balance(connection, customer_id)

        cursor.execute("""
            SELECT id, transaction_date, transaction_type,
                   amount, note, created_at
            FROM loan_transactions
            WHERE customer_id = %s
            ORDER BY transaction_date DESC, id DESC
        """, (customer_id,))
        transactions = cursor.fetchall()

    except Error as error:
        flash(f"Database error: {error}", "danger")
        return redirect(url_for("customer_loans"))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="grid">
        <div class="card">
            <div class="stat-title">Customer</div>
            <div class="stat-value" style="font-size:23px;">
                {{ customer.customer_name }}
            </div>
            <p>{{ customer.customer_code }} · {{ customer.mobile_number }}</p>
        </div>

        <div class="card">
            <div class="stat-title">Current Outstanding Balance</div>
            <div class="stat-value {{ 'balance-positive' if balance > 0 else 'balance-zero' }}">
                Rs. {{ balance|money }}
            </div>
        </div>

        <div class="card">
            <div class="stat-title">Number of Transactions</div>
            <div class="stat-value">{{ transactions|length }}</div>
        </div>
    </div>

    <div class="card">
        <div class="section-header">
            <h2>Add Loan or Payment</h2>
            <div class="action-buttons">
                <a class="btn btn-secondary" href="{{ url_for('edit_customer', customer_id=customer.id) }}">Edit Customer</a>
                <a class="btn btn-secondary" href="{{ url_for('customer_loans') }}">Back to Customers</a>
            </div>
        </div>

        <form method="POST"
              action="{{ url_for('add_transaction', customer_id=customer.id) }}">
            <div class="form-grid">
                <div class="form-group">
                    <label>Transaction Type</label>
                    <select name="transaction_type" required>
                        <option value="LOAN">New Loan / Credit Purchase</option>
                        <option value="PAYMENT">Customer Payment</option>
                    </select>
                </div>

                <div class="form-group">
                    <label>Date</label>
                    <input type="date" name="transaction_date"
                           value="{{ today }}" required>
                </div>

                <div class="form-group">
                    <label>Amount (Rs.)</label>
                    <input type="number" name="amount" min="0.01"
                           step="0.01" required>
                </div>

                <div class="form-group">
                    <label>Note</label>
                    <input type="text" name="note"
                           placeholder="Example: Rice and groceries">
                </div>

                <div class="full">
                    <button class="btn btn-primary" type="submit">
                        Save Transaction
                    </button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>{{ customer.customer_name }} - History</h2>

        {% if transactions %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Type</th>
                        <th>Loan Added</th>
                        <th>Payment</th>
                        <th>Note</th>
                        <th>Recorded At</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                {% for transaction in transactions %}
                    <tr>
                        <td>{{ transaction.transaction_date }}</td>
                        <td>
                            <span class="badge {{ 'badge-loan' if transaction.transaction_type == 'LOAN' else 'badge-payment' }}">
                                {{ transaction.transaction_type }}
                            </span>
                        </td>
                        <td>
                            {% if transaction.transaction_type == 'LOAN' %}
                                Rs. {{ transaction.amount|money }}
                            {% else %}-{% endif %}
                        </td>
                        <td>
                            {% if transaction.transaction_type == 'PAYMENT' %}
                                Rs. {{ transaction.amount|money }}
                            {% else %}-{% endif %}
                        </td>
                        <td>{{ transaction.note or '-' }}</td>
                        <td>{{ transaction.created_at|sl_time }}</td>
                        <td>
                            <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
                                <a class="btn btn-secondary"
                                   href="{{ url_for('edit_transaction', transaction_id=transaction.id) }}">
                                    Edit
                                </a>

                                <form method="POST"
                                      action="{{ url_for('delete_transaction', transaction_id=transaction.id) }}"
                                      onsubmit="return confirmDelete('Delete this transaction?');"
                                      style="margin:0;">
                                    <button class="btn btn-danger" type="submit">
                                        <svg viewBox="0 0 24 24" fill="none" stroke-width="2" aria-hidden="true">
                                            <path d="M3 6h18"/><path d="M8 6V4h8v2"/>
                                            <path d="M19 6l-1 14H6L5 6"/><path d="M10 11v5M14 11v5"/>
                                        </svg>
                                        Delete
                                    </button>
                                </form>
                            </div>
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No transactions for this customer.</div>
        {% endif %}
    </div>
    """

    return render_page(
        "Customer Details",
        customer["customer_name"],
        template,
        active_page="loans",
        customer=customer,
        balance=balance,
        transactions=transactions,
        today=date.today().isoformat()
    )


@app.route(
    "/customer-loans/customer/<int:customer_id>/transaction",
    methods=["POST"]
)
@login_required
def add_transaction(customer_id):
    transaction_type = request.form.get("transaction_type", "").upper()
    transaction_date = request.form.get("transaction_date", "")
    amount_text = request.form.get("amount", "").strip()
    note = request.form.get("note", "").strip()

    if transaction_type not in ("LOAN", "PAYMENT"):
        flash("Invalid transaction type.", "danger")
        return redirect(url_for("customer_details", customer_id=customer_id))

    try:
        amount = Decimal(amount_text)
        if amount <= 0:
            raise InvalidOperation
    except (InvalidOperation, ValueError):
        flash("Amount must be greater than zero.", "danger")
        return redirect(url_for("customer_details", customer_id=customer_id))

    connection = None
    cursor = None
    try:
        connection = db_connection()

        if transaction_type == "PAYMENT":
            current_balance = get_customer_balance(connection, customer_id)
            if amount > current_balance:
                flash(
                    f"Payment cannot be greater than the current balance "
                    f"(Rs. {money(current_balance)}).",
                    "danger"
                )
                return redirect(
                    url_for("customer_details", customer_id=customer_id)
                )

        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO loan_transactions
                (customer_id, transaction_date,
                 transaction_type, amount, note)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            customer_id,
            transaction_date or date.today().isoformat(),
            transaction_type,
            amount,
            note or None
        ))
        connection.commit()

        new_balance = get_customer_balance(connection, customer_id)
        flash(
            f"Transaction saved. Current balance: Rs. {money(new_balance)}",
            "success"
        )

    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    return redirect(url_for("customer_details", customer_id=customer_id))


@app.route(
    "/customer-loans/transaction/<int:transaction_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_transaction(transaction_id):
    connection = None
    cursor = None

    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, customer_id, transaction_date,
                   transaction_type, amount, note
            FROM loan_transactions
            WHERE id = %s
        """, (transaction_id,))
        transaction = cursor.fetchone()

        if not transaction:
            flash("Transaction not found.", "danger")
            return redirect(url_for("customer_loans"))

        customer_id = transaction["customer_id"]

        if request.method == "POST":
            transaction_type = request.form.get("transaction_type", "").upper()
            transaction_date = request.form.get("transaction_date", "")
            amount_text = request.form.get("amount", "").strip()
            note = request.form.get("note", "").strip()

            if transaction_type not in ("LOAN", "PAYMENT"):
                flash("Invalid transaction type.", "danger")
            elif not transaction_date:
                flash("Transaction date is required.", "danger")
            else:
                try:
                    amount = Decimal(amount_text)
                    if amount <= 0:
                        raise InvalidOperation

                    if transaction_type == "PAYMENT":
                        cursor.execute("""
                            SELECT COALESCE(SUM(
                                CASE
                                    WHEN transaction_type = 'LOAN' THEN amount
                                    WHEN transaction_type = 'PAYMENT' THEN -amount
                                    ELSE 0
                                END
                            ), 0) AS balance_without_current
                            FROM loan_transactions
                            WHERE customer_id = %s AND id <> %s
                        """, (customer_id, transaction_id))
                        balance_without_current = Decimal(
                            cursor.fetchone()["balance_without_current"] or 0
                        )

                        if amount > balance_without_current:
                            flash(
                                f"Payment cannot be greater than the available balance "
                                f"(Rs. {money(balance_without_current)}).",
                                "danger"
                            )
                            return redirect(
                                url_for("edit_transaction", transaction_id=transaction_id)
                            )

                    cursor.execute("""
                        UPDATE loan_transactions
                        SET transaction_date = %s,
                            transaction_type = %s,
                            amount = %s,
                            note = %s
                        WHERE id = %s
                    """, (
                        transaction_date,
                        transaction_type,
                        amount,
                        note or None,
                        transaction_id
                    ))
                    connection.commit()

                    new_balance = get_customer_balance(connection, customer_id)
                    flash(
                        f"Transaction updated successfully. Current balance: "
                        f"Rs. {money(new_balance)}",
                        "success"
                    )
                    return redirect(
                        url_for("customer_details", customer_id=customer_id)
                    )

                except (InvalidOperation, ValueError):
                    flash("Amount must be greater than zero.", "danger")

        template = r"""
        <div class="card">
            <div class="section-header">
                <h2>Edit Transaction</h2>
                <a class="btn btn-secondary"
                   href="{{ url_for('customer_details', customer_id=transaction.customer_id) }}">
                    Cancel
                </a>
            </div>

            <form method="POST">
                <div class="form-grid">
                    <div class="form-group">
                        <label>Transaction Type</label>
                        <select name="transaction_type" required>
                            <option value="LOAN"
                                {{ 'selected' if transaction.transaction_type == 'LOAN' else '' }}>
                                New Loan / Credit Purchase
                            </option>
                            <option value="PAYMENT"
                                {{ 'selected' if transaction.transaction_type == 'PAYMENT' else '' }}>
                                Customer Payment
                            </option>
                        </select>
                    </div>

                    <div class="form-group">
                        <label>Date</label>
                        <input type="date" name="transaction_date"
                               value="{{ transaction.transaction_date }}" required>
                    </div>

                    <div class="form-group">
                        <label>Amount (Rs.)</label>
                        <input type="number" name="amount" min="0.01" step="0.01"
                               value="{{ transaction.amount }}" required>
                    </div>

                    <div class="form-group">
                        <label>Note</label>
                        <input type="text" name="note"
                               value="{{ transaction.note or '' }}"
                               placeholder="Example: Rice and groceries">
                    </div>

                    <div class="full">
                        <button class="btn btn-primary" type="submit">
                            Update Transaction
                        </button>
                    </div>
                </div>
            </form>
        </div>
        """

        return render_page(
            "Edit Transaction",
            "Edit Transaction",
            template,
            active_page="loans",
            transaction=transaction
        )

    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
        return redirect(url_for("customer_loans"))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route(
    "/customer-loans/transaction/<int:transaction_id>/delete",
    methods=["POST"]
)
@login_required
def delete_transaction(transaction_id):
    connection = None
    cursor = None
    customer_id = None

    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            "SELECT customer_id FROM loan_transactions WHERE id = %s",
            (transaction_id,)
        )
        transaction = cursor.fetchone()

        if not transaction:
            flash("Transaction not found.", "danger")
            return redirect(url_for("customer_loans"))

        customer_id = transaction["customer_id"]
        cursor.execute(
            "DELETE FROM loan_transactions WHERE id = %s",
            (transaction_id,)
        )
        connection.commit()
        flash("Transaction deleted.", "success")

    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    if customer_id:
        return redirect(url_for("customer_details", customer_id=customer_id))
    return redirect(url_for("customer_loans"))


# ------------------------- SHOP ITEMS -------------------------

@app.route("/shop-items", methods=["GET", "POST"])
@login_required
def shop_items():
    if request.method == "POST":
        item_name = request.form.get("item_name", "").strip()
        item_code = request.form.get("item_code", "").strip() or None
        quantity = request.form.get("quantity", "0").strip()
        description = request.form.get("description", "").strip()

        connection = None
        cursor = None
        try:
            quantity_value = int(quantity)
            if quantity_value < 0:
                raise ValueError

            if not item_name:
                flash("Item name is required.", "danger")
            else:
                connection = db_connection()
                cursor = connection.cursor()
                cursor.execute("""
                    INSERT INTO shop_items
                        (item_name, item_code, quantity, description)
                    VALUES (%s, %s, %s, %s)
                """, (
                    item_name,
                    item_code,
                    quantity_value,
                    description or None
                ))
                connection.commit()
                flash("Shop item saved successfully.", "success")
                return redirect(url_for("shop_items"))

        except ValueError:
            flash("Quantity must be zero or a positive whole number.", "danger")
        except Error as error:
            if connection:
                connection.rollback()
            if error.errno == 1062:
                flash("That item code already exists.", "danger")
            else:
                flash(f"Database error: {error}", "danger")
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()

    connection = None
    cursor = None
    items = []
    weekly_added = 0
    weekly_finished = 0
    active_items = 0
    week_start = date.today() - timedelta(days=date.today().weekday())
    week_end = week_start + timedelta(days=6)
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM shop_items ORDER BY id DESC LIMIT 300")
        items = cursor.fetchall()

        # Calculate all item counters in a single query instead of three
        # separate network round-trips.
        cursor.execute("""
            SELECT
                SUM(created_at >= %s AND created_at < DATE_ADD(%s, INTERVAL 1 DAY)) AS weekly_added,
                SUM(finished_at IS NOT NULL
                    AND finished_at >= %s
                    AND finished_at < DATE_ADD(%s, INTERVAL 1 DAY)) AS weekly_finished,
                SUM(finished_at IS NULL) AS active_items
            FROM shop_items
        """, (week_start, week_end, week_start, week_end))
        item_stats = cursor.fetchone() or {}
        weekly_added = int(item_stats.get("weekly_added") or 0)
        weekly_finished = int(item_stats.get("weekly_finished") or 0)
        active_items = int(item_stats.get("active_items") or 0)
    except Error as error:
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="grid">
        <div class="card">
            <div class="stat-title">Items Added This Week</div>
            <div class="stat-value">{{ weekly_added }}</div>
            <small style="color:var(--muted);">{{ week_start }} to {{ week_end }}</small>
        </div>
        <div class="card">
            <div class="stat-title">Items Finished This Week</div>
            <div class="stat-value" style="color:var(--success);">{{ weekly_finished }}</div>
            <small style="color:var(--muted);">Marked as finished during this week</small>
        </div>
        <div class="card">
            <div class="stat-title">Still Active</div>
            <div class="stat-value">{{ active_items }}</div>
            <small style="color:var(--muted);">Items not finished yet</small>
        </div>
    </div>

    <div class="card">
        <h2>Add Shop Thing</h2>
        <form method="POST">
            <div class="form-grid">
                <div class="form-group">
                    <label>Item Name</label>
                    <input type="text" name="item_name" required>
                </div>

                <div class="form-group">
                    <label>Item Code</label>
                    <input type="text" name="item_code"
                           placeholder="Optional unique code">
                </div>

                <div class="form-group">
                    <label>Quantity</label>
                    <input type="number" name="quantity" min="0" value="0" required>
                </div>

                <div class="form-group">
                    <label>Description</label>
                    <input type="text" name="description">
                </div>

                <div class="full">
                    <button class="btn btn-primary" type="submit">Save Item</button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Shop Things History</h2>
        {% if items %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Item Code</th>
                        <th>Item Name</th>
                        <th>Quantity</th>
                        <th>Description</th>
                        <th>Status</th>
                        <th>Created</th>
                        <th>Finished</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                {% for item in items %}
                    <tr>
                        <td>{{ item.item_code or '-' }}</td>
                        <td>{{ item.item_name }}</td>
                        <td>{{ item.quantity }}</td>
                        <td>{{ item.description or '-' }}</td>
                        <td>
                            {% if item.finished_at %}
                                <span class="badge badge-payment">Finished</span>
                            {% else %}
                                <span class="badge badge-loan">Active</span>
                            {% endif %}
                        </td>
                        <td>{{ item.created_at|sl_time }}</td>
                        <td>{{ item.finished_at|sl_time if item.finished_at else '-' }}</td>
                        <td>
                            <div class="action-buttons">
                                <a class="btn btn-secondary" href="{{ url_for('edit_item', item_id=item.id) }}">Edit</a>
                                {% if item.finished_at %}
                                <form method="POST" action="{{ url_for('reopen_item', item_id=item.id) }}">
                                    <button class="btn btn-secondary" type="submit">Reopen</button>
                                </form>
                                {% else %}
                                <form method="POST" action="{{ url_for('finish_item', item_id=item.id) }}">
                                    <button class="btn btn-success" type="submit">Finish</button>
                                </form>
                                {% endif %}
                                <form method="POST"
                                      action="{{ url_for('delete_item', item_id=item.id) }}"
                                      onsubmit="return confirmDelete('Delete this item?');">
                                    <button class="btn btn-danger" type="submit">Delete</button>
                                </form>
                            </div>
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No shop things have been added.</div>
        {% endif %}
    </div>
    """

    return render_page(
        "Shop Things",
        "Shop Things Details",
        template,
        active_page="items",
        items=items,
        weekly_added=weekly_added,
        weekly_finished=weekly_finished,
        active_items=active_items,
        week_start=week_start.strftime("%d %b %Y"),
        week_end=week_end.strftime("%d %b %Y")
    )


@app.route("/shop-items/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit_item(item_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM shop_items WHERE id = %s", (item_id,))
        item = cursor.fetchone()
        if not item:
            flash("Shop item not found.", "danger")
            return redirect(url_for("shop_items"))

        if request.method == "POST":
            item_name = request.form.get("item_name", "").strip()
            item_code = request.form.get("item_code", "").strip() or None
            quantity_text = request.form.get("quantity", "0").strip()
            description = request.form.get("description", "").strip()
            try:
                quantity = int(quantity_text)
                if quantity < 0:
                    raise ValueError
                if not item_name:
                    flash("Item name is required.", "danger")
                else:
                    cursor.execute("""
                        UPDATE shop_items
                        SET item_name = %s, item_code = %s, quantity = %s, description = %s
                        WHERE id = %s
                    """, (item_name, item_code, quantity, description or None, item_id))
                    connection.commit()
                    flash("Shop item updated successfully.", "success")
                    return redirect(url_for("shop_items"))
            except ValueError:
                flash("Quantity must be zero or a positive whole number.", "danger")
            except Error as error:
                connection.rollback()
                if error.errno == 1062:
                    flash("That item code already exists.", "danger")
                else:
                    raise

        template = r"""
        <div class="card">
            <div class="section-header">
                <h2>Edit Shop Thing</h2>
                <a class="btn btn-secondary" href="{{ url_for('shop_items') }}">Cancel</a>
            </div>
            <form method="POST">
                <div class="form-grid">
                    <div class="form-group"><label>Item Name</label><input type="text" name="item_name" value="{{ item.item_name }}" required></div>
                    <div class="form-group"><label>Item Code</label><input type="text" name="item_code" value="{{ item.item_code or '' }}"></div>
                    <div class="form-group"><label>Quantity</label><input type="number" name="quantity" min="0" value="{{ item.quantity }}" required></div>
                    <div class="form-group"><label>Description</label><input type="text" name="description" value="{{ item.description or '' }}"></div>
                    <div class="full"><button class="btn btn-primary" type="submit">Update Item</button></div>
                </div>
            </form>
        </div>
        """
        return render_page("Edit Shop Thing", "Edit Shop Thing", template,
                           active_page="items", item=item)
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
        return redirect(url_for("shop_items"))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route("/shop-items/<int:item_id>/finish", methods=["POST"])
@login_required
def finish_item(item_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE shop_items
            SET finished_at = NOW()
            WHERE id = %s AND finished_at IS NULL
        """, (item_id,))
        connection.commit()
        flash("Item marked as finished.", "success")
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()
    return redirect(url_for("shop_items"))


@app.route("/shop-items/<int:item_id>/reopen", methods=["POST"])
@login_required
def reopen_item(item_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute("UPDATE shop_items SET finished_at = NULL WHERE id = %s", (item_id,))
        connection.commit()
        flash("Item reopened.", "success")
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()
    return redirect(url_for("shop_items"))


@app.route("/shop-items/<int:item_id>/delete", methods=["POST"])
@login_required
def delete_item(item_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute("DELETE FROM shop_items WHERE id = %s", (item_id,))
        connection.commit()
        flash("Shop item deleted.", "success")
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    return redirect(url_for("shop_items"))


# ---------------------- PRICE MANAGEMENT ----------------------

@app.route("/prices", methods=["GET", "POST"])
@login_required
def price_management():
    if request.method == "POST":
        item_name = request.form.get("item_name", "").strip()
        buying_text = request.form.get("buying_price", "").strip()
        selling_text = request.form.get("selling_price", "").strip()

        connection = None
        cursor = None
        try:
            buying_price = Decimal(buying_text)
            selling_price = Decimal(selling_text)
            if buying_price < 0 or selling_price < 0:
                raise InvalidOperation

            if not item_name:
                flash("Item name is required.", "danger")
            else:
                connection = db_connection()
                cursor = connection.cursor()
                cursor.execute("""
                    INSERT INTO prices (item_name, buying_price, selling_price)
                    VALUES (%s, %s, %s)
                """, (item_name, buying_price, selling_price))
                connection.commit()
                flash("Item price saved successfully.", "success")
                return redirect(url_for("price_management"))

        except (InvalidOperation, ValueError):
            flash("Buying price and selling price must be valid positive numbers.", "danger")
        except Error as error:
            if connection:
                connection.rollback()
            flash(f"Database error: {error}", "danger")
        finally:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()

    connection = None
    cursor = None
    prices = []
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM prices ORDER BY id DESC LIMIT 300")
        prices = cursor.fetchall()
    except Error as error:
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="card">
        <h2>Add Item Price</h2>
        <form method="POST">
            <div class="form-grid">
                <div class="form-group">
                    <label>Item Name</label>
                    <input type="text" name="item_name" required>
                </div>

                <div class="form-group">
                    <label>Buying Price (Rs.)</label>
                    <input type="number" name="buying_price"
                           min="0" step="0.01" required>
                </div>

                <div class="form-group">
                    <label>Selling Price (Rs.)</label>
                    <input type="number" name="selling_price"
                           min="0" step="0.01" required>
                </div>

                <div class="full">
                    <button class="btn btn-primary" type="submit">Save Price</button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Price List</h2>
        {% if prices %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Item Name</th>
                        <th>Buying Price</th>
                        <th>Selling Price</th>
                        <th>Profit / Item</th>
                        <th>Last Updated</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                {% for price in prices %}
                    <tr>
                        <td>{{ price.item_name }}</td>
                        <td>Rs. {{ price.buying_price|money }}</td>
                        <td>Rs. {{ price.selling_price|money }}</td>
                        <td>Rs. {{ (price.selling_price - price.buying_price)|money }}</td>
                        <td>{{ price.updated_at|sl_time }}</td>
                        <td>
                            <div class="action-buttons">
                            <a class="btn btn-secondary" href="{{ url_for('edit_price', price_id=price.id) }}">Edit</a>
                            <form method="POST"
                                  action="{{ url_for('delete_price', price_id=price.id) }}"
                                  onsubmit="return confirmDelete('Delete this price?');">
                                <button class="btn btn-danger" type="submit">Delete</button>
                            </form>
                            </div>
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No prices have been added.</div>
        {% endif %}
    </div>
    """

    return render_page(
        "Prices",
        "Price Management",
        template,
        active_page="prices",
        prices=prices
    )


@app.route("/prices/<int:price_id>/edit", methods=["GET", "POST"])
@login_required
def edit_price(price_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM prices WHERE id = %s", (price_id,))
        price = cursor.fetchone()
        if not price:
            flash("Price record not found.", "danger")
            return redirect(url_for("price_management"))

        if request.method == "POST":
            item_name = request.form.get("item_name", "").strip()
            buying_text = request.form.get("buying_price", "").strip()
            selling_text = request.form.get("selling_price", "").strip()
            try:
                buying_price = Decimal(buying_text)
                selling_price = Decimal(selling_text)
                if buying_price < 0 or selling_price < 0:
                    raise InvalidOperation
                if not item_name:
                    flash("Item name is required.", "danger")
                else:
                    cursor.execute("""
                        UPDATE prices
                        SET item_name = %s, buying_price = %s, selling_price = %s
                        WHERE id = %s
                    """, (item_name, buying_price, selling_price, price_id))
                    connection.commit()
                    flash("Price updated successfully.", "success")
                    return redirect(url_for("price_management"))
            except (InvalidOperation, ValueError):
                flash("Buying price and selling price must be valid positive numbers.", "danger")

        template = r"""
        <div class="card">
            <div class="section-header">
                <h2>Edit Item Price</h2>
                <a class="btn btn-secondary" href="{{ url_for('price_management') }}">Cancel</a>
            </div>
            <form method="POST">
                <div class="form-grid">
                    <div class="form-group"><label>Item Name</label><input type="text" name="item_name" value="{{ price.item_name }}" required></div>
                    <div class="form-group"><label>Buying Price (Rs.)</label><input type="number" name="buying_price" min="0" step="0.01" value="{{ price.buying_price }}" required></div>
                    <div class="form-group"><label>Selling Price (Rs.)</label><input type="number" name="selling_price" min="0" step="0.01" value="{{ price.selling_price }}" required></div>
                    <div class="full"><button class="btn btn-primary" type="submit">Update Price</button></div>
                </div>
            </form>
        </div>
        """
        return render_page("Edit Price", "Edit Item Price", template,
                           active_page="prices", price=price)
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
        return redirect(url_for("price_management"))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()


@app.route("/prices/<int:price_id>/delete", methods=["POST"])
@login_required
def delete_price(price_id):
    connection = None
    cursor = None
    try:
        connection = db_connection()
        cursor = connection.cursor()
        cursor.execute("DELETE FROM prices WHERE id = %s", (price_id,))
        connection.commit()
        flash("Price deleted.", "success")
    except Error as error:
        if connection:
            connection.rollback()
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    return redirect(url_for("price_management"))



# ------------------------- LIVE LOGS --------------------------

@app.route("/live-logs")
@login_required
def live_logs():
    """Phone-friendly live device and activity monitor."""
    connection = None
    cursor = None
    recent_logs = []
    devices = []
    online_count = 0

    try:
        connection = db_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, username, activity, endpoint, method, ip_address,
                   device_type, browser, platform, created_at
            FROM activity_logs
            ORDER BY created_at DESC, id DESC
            LIMIT 250
        """)
        recent_logs = cursor.fetchall()

        cursor.execute("""
            SELECT
                username, ip_address, device_type, browser, platform, user_agent,
                MAX(created_at) AS last_seen,
                SUBSTRING_INDEX(
                    GROUP_CONCAT(activity ORDER BY created_at DESC, id DESC SEPARATOR '|||'),
                    '|||', 1
                ) AS last_activity,
                SUBSTRING_INDEX(
                    GROUP_CONCAT(endpoint ORDER BY created_at DESC, id DESC SEPARATOR '|||'),
                    '|||', 1
                ) AS last_endpoint
            FROM activity_logs
            WHERE username IS NOT NULL
              AND created_at >= (NOW() - INTERVAL 30 DAY)
            GROUP BY username, ip_address, device_type, browser, platform, user_agent
            ORDER BY last_seen DESC
            LIMIT 60
        """)
        devices = cursor.fetchall()

        # Keep online/offline comparisons in UTC. Display conversion happens
        # separately so adding +05:30 cannot affect the five-minute status window.
        now_utc = datetime.now(timezone.utc)
        for d in devices:
            seen = d.get("last_seen")
            if seen:
                seen_utc = (
                    seen.replace(tzinfo=timezone.utc)
                    if seen.tzinfo is None
                    else seen.astimezone(timezone.utc)
                )
                d["online"] = (now_utc - seen_utc) <= timedelta(minutes=5)
                d["last_seen_display"] = to_sri_lanka_time(seen)
            else:
                d["online"] = False
                d["last_seen_display"] = "-"

            if d["online"]:
                online_count += 1

        for log in recent_logs:
            log["created_at_display"] = to_sri_lanka_time(log.get("created_at"))

    except Error as error:
        flash(f"Database error: {error}", "danger")
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

    template = r"""
    <div class="data-hero">
        <div class="data-hero-main">
            <h2><span class="logs-live-dot"></span>&nbsp; Live Device & Activity Logs</h2>
            <p>See which devices are using Ambaal Shop, their browser, IP address, recent activity and last-seen time. The page refreshes automatically every 15 seconds.</p>
        </div>
        <div class="data-db-card">
            <small>Devices online now</small>
            <strong style="font-size:30px;color:var(--success);">{{ online_count }}</strong>
            <small style="margin-top:10px;">Online = activity within 5 minutes</small>
        </div>
    </div>

    <div class="logs-toolbar">
        <a class="btn btn-primary" href="{{ url_for('live_logs') }}" title="Refresh logs">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M20 11a8.1 8.1 0 1 0 2 5.3"/><path d="M20 4v7h-7"/></svg>
            Refresh
        </a>
        <span style="color:var(--muted);font-size:12px;font-weight:700;">Recent history: up to 250 activities · stored for 30 days</span>
    </div>

    {% if devices %}
    <div class="device-grid">
        {% for d in devices %}
        <div class="device-card">
            <div class="device-card-head">
                <div class="device-name">
                    {% if 'iPhone' in d.device_type %}📱{% elif 'Android' in d.device_type %}📱{% elif 'iPad' in d.device_type or 'Tablet' in d.device_type %}▣{% else %}💻{% endif %}
                    {{ d.device_type }}
                </div>
                <span class="log-status {{ 'status-online' if d.online else 'status-offline' }}">
                    {{ '● Online' if d.online else '○ Offline' }}
                </span>
            </div>
            <div class="device-meta">
                <strong>User:</strong> {{ d.username }}<br>
                <strong>Browser:</strong> {{ d.browser }} · {{ d.platform }}<br>
                <strong>IP:</strong> {{ d.ip_address or 'Unknown' }}<br>
                <strong>Last activity:</strong> {{ d.last_activity or '-' }}<br>
                <strong>Last seen:</strong> {{ d.last_seen_display }} <small style="color:var(--muted);">(Sri Lanka)</small>
            </div>
        </div>
        {% endfor %}
    </div>
    {% else %}
        <div class="empty">No device activity has been recorded yet. Log in from a device and use the system to create the first record.</div>
    {% endif %}

    <div class="card data-section">
        <div class="data-section-head">
            <h2>Recent Activity</h2>
            <span class="data-count">{{ recent_logs|length }}</span>
        </div>
        {% if recent_logs %}
        <div class="table-wrap">
            <table class="log-table">
                <thead><tr>
                    <th>#</th><th>Time</th><th>User</th><th>Device</th><th>Browser</th>
                    <th>IP Address</th><th>Activity</th><th>Method</th>
                </tr></thead>
                <tbody>
                {% for x in recent_logs %}
                <tr>
                    <td>{{ x.id }}</td>
                    <td>{{ x.created_at_display }}<br><small style="color:var(--muted);">Sri Lanka time</small></td>
                    <td><strong>{{ x.username or '-' }}</strong></td>
                    <td>{{ x.device_type or '-' }}<br><small style="color:var(--muted);">{{ x.platform or '' }}</small></td>
                    <td>{{ x.browser or '-' }}</td>
                    <td class="log-ip">{{ x.ip_address or '-' }}</td>
                    <td class="log-action">{{ x.activity }}</td>
                    <td><span class="badge">{{ x.method or '-' }}</span></td>
                </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <div class="empty">No activity logs found.</div>
        {% endif %}
    </div>

    <script>
        setTimeout(() => window.location.reload(), 15000);
    </script>
    """

    return render_page(
        "Live Logs",
        "Live Logs",
        template,
        active_page="logs",
        recent_logs=recent_logs,
        devices=devices,
        online_count=online_count
    )


# ---------------------- ALL DATA VIEWER -----------------------


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _style_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top")
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 34)


def _load_all_data():
    connection = db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT c.id, c.customer_code, c.customer_name, c.mobile_number, c.created_at,
                   COALESCE(SUM(CASE
                       WHEN lt.transaction_type = 'LOAN' THEN lt.amount
                       WHEN lt.transaction_type = 'PAYMENT' THEN -lt.amount
                       ELSE 0 END), 0) AS balance
            FROM customers c
            LEFT JOIN loan_transactions lt ON lt.customer_id = c.id
            GROUP BY c.id, c.customer_code, c.customer_name, c.mobile_number, c.created_at
            ORDER BY c.customer_name
        """)
        customers = cursor.fetchall()

        cursor.execute("""
            SELECT lt.id, lt.transaction_date, lt.transaction_type, lt.amount,
                   lt.note, lt.created_at, c.customer_code, c.customer_name
            FROM loan_transactions lt
            JOIN customers c ON c.id = lt.customer_id
            ORDER BY lt.transaction_date DESC, lt.id DESC
        """)
        transactions = cursor.fetchall()
        cursor.execute("SELECT id, item_code, item_name, quantity, description, finished_at, created_at FROM shop_items ORDER BY id DESC")
        items = cursor.fetchall()
        cursor.execute("SELECT id, item_name, buying_price, selling_price, updated_at FROM prices ORDER BY id DESC")
        prices = cursor.fetchall()
        cursor.execute("SELECT id, username, created_at FROM users ORDER BY id")
        users = cursor.fetchall()
        return customers, transactions, items, prices, users
    finally:
        cursor.close()
        connection.close()


@app.route("/all-data/export-excel")
@login_required
def export_all_data_excel():
    """Download the data store as a real .xlsx Excel workbook."""
    sheet = request.args.get("sheet", "all").strip().lower()
    customers, transactions, items, prices, users = _load_all_data()

    wb = Workbook()
    wb.remove(wb.active)

    datasets = {
        "customers": ("Customers", ["No.", "Customer ID", "Name", "Mobile", "Balance (Rs.)", "Created"],
                      [[i, x["customer_code"], x["customer_name"], x["mobile_number"], _excel_value(x["balance"]), x["created_at"]] for i, x in enumerate(customers, 1)]),
        "transactions": ("Transactions", ["No.", "Date", "Customer", "Customer ID", "Type", "Amount (Rs.)", "Note", "Recorded"],
                         [[i, x["transaction_date"], x["customer_name"], x["customer_code"], x["transaction_type"], _excel_value(x["amount"]), x["note"] or "", x["created_at"]] for i, x in enumerate(transactions, 1)]),
        "items": ("Shop Items", ["No.", "Code", "Item", "Quantity", "Description", "Created"],
                  [[i, x["item_code"] or "", x["item_name"], x["quantity"], x["description"] or "", x["created_at"]] for i, x in enumerate(items, 1)]),
        "prices": ("Prices", ["No.", "Item", "Selling Price (Rs.)", "Updated"],
                   [[i, x["item_name"], _excel_value(x["selling_price"]), x["updated_at"]] for i, x in enumerate(prices, 1)]),
        "users": ("System Users", ["No.", "User ID", "Username", "Created"],
                  [[i, x["id"], x["username"], x["created_at"]] for i, x in enumerate(users, 1)])
    }

    selected = datasets.items() if sheet == "all" else [(sheet, datasets[sheet])] if sheet in datasets else datasets.items()
    for _, (title, headers, rows) in selected:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for row in rows:
            ws.append(row)
        _style_excel_sheet(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ambaal_shop_{sheet}_data.xlsx" if sheet != "all" else "ambaal_shop_all_data.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/all-data")
@login_required
def all_data():
    """Spreadsheet-style view of the main application data."""
    search = request.args.get("search", "").strip()
    customers = transactions = items = prices = users = []
    try:
        customers, transactions, items, prices, users = _load_all_data()
        if search:
            q = search.lower()
            def contains(*values):
                return any(q in str(v or "").lower() for v in values)
            customers = [x for x in customers if contains(x["customer_code"], x["customer_name"], x["mobile_number"], x["balance"])]
            transactions = [x for x in transactions if contains(x["customer_code"], x["customer_name"], x["transaction_type"], x["amount"], x["note"], x["transaction_date"])]
            items = [x for x in items if contains(x["item_code"], x["item_name"], x["quantity"], x["description"])]
            prices = [x for x in prices if contains(x["item_name"], x["buying_price"], x["selling_price"])]
            users = [x for x in users if contains(x["username"])]
    except Error as error:
        flash(f"Database error: {error}", "danger")

    template = r"""
    <div class="data-hero">
        <div class="data-hero-main">
            <h2>Data Store</h2>
            <p>All database records are displayed in spreadsheet-style tables. Scroll horizontally on smaller devices and use the Excel buttons to save .xlsx files directly to your device.</p>
        </div>
        <div class="data-db-card">
            <small>Connected database</small><strong>{{ db_name }}</strong>
            <small style="margin-top:12px;">Server</small><strong style="font-size:14px;">{{ db_host }}:{{ db_port }}</strong>
        </div>
    </div>

    <div class="excel-toolbar">
        <a class="btn excel-btn" href="{{ url_for('export_all_data_excel', sheet='all') }}">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg>
            Download All Excel
        </a>
        <form class="data-search" method="GET" style="margin:0; flex:1;">
            <input type="search" name="search" value="{{ search }}" placeholder="Search all data...">
            <button class="btn btn-primary" type="submit">
                <svg viewBox="0 0 24 24" fill="none" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="m20 20-3.5-3.5"/></svg> Search
            </button>
            {% if search %}<a class="btn btn-secondary" href="{{ url_for('all_data') }}">Clear</a>{% endif %}
        </form>
    </div>

    <div class="privacy-note">Security: password hashes remain hidden. Excel exports contain only the same safe fields displayed on this page.</div>

    <div class="excel-sheet">
      <div class="excel-sheet-head"><div class="excel-sheet-title"><h2>Customers</h2><span class="data-count">{{ customers|length }}</span></div><a class="icon-btn download" title="Download Customers Excel" href="{{ url_for('export_all_data_excel', sheet='customers') }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg></a></div>
      {% if customers %}<div class="excel-grid-wrap"><table class="excel-table"><thead><tr><th>#</th><th>Customer ID</th><th>Name</th><th>Mobile</th><th>Balance</th><th>Created</th><th>Actions</th></tr></thead><tbody>
      {% for x in customers %}<tr><td class="row-no">{{ loop.index }}</td><td>{{ x.customer_code }}</td><td>{{ x.customer_name }}</td><td>{{ x.mobile_number }}</td><td class="{{ 'balance-positive' if x.balance > 0 else 'balance-zero' }}">Rs. {{ x.balance|money }}</td><td>{{ x.created_at|sl_time }}</td><td><div class="excel-actions"><a class="icon-btn view" title="View Customer" href="{{ url_for('customer_details', customer_id=x.id) }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M2 12s3.5-7 10-7 10 7 10 7-3.5 7-10 7S2 12 2 12Z"/><circle cx="12" cy="12" r="3"/></svg></a><a class="icon-btn edit" title="Edit Customer" href="{{ url_for('edit_customer', customer_id=x.id) }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 20h9"/><path d="M16.5 3.5a2.1 2.1 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg></a><form method="POST" action="{{ url_for('delete_customer', customer_id=x.id) }}" onsubmit="return confirm('Delete this customer and all related transactions?');"><button class="icon-btn delete" title="Delete Customer" type="submit"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 6h18"/><path d="M8 6V4h8v2"/><path d="M19 6l-1 14H6L5 6"/><path d="M10 11v5M14 11v5"/></svg></button></form></div></td></tr>{% endfor %}
      </tbody></table></div>{% else %}<div class="empty">No customer data found.</div>{% endif %}
    </div>

    <div class="excel-sheet">
      <div class="excel-sheet-head"><div class="excel-sheet-title"><h2>Loan & Payment Transactions</h2><span class="data-count">{{ transactions|length }}</span></div><a class="icon-btn download" title="Download Transactions Excel" href="{{ url_for('export_all_data_excel', sheet='transactions') }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg></a></div>
      {% if transactions %}<div class="excel-grid-wrap"><table class="excel-table"><thead><tr><th>#</th><th>Date</th><th>Customer</th><th>ID</th><th>Type</th><th>Amount</th><th>Note</th><th>Recorded</th><th>Actions</th></tr></thead><tbody>
      {% for x in transactions %}<tr><td class="row-no">{{ loop.index }}</td><td>{{ x.transaction_date }}</td><td>{{ x.customer_name }}</td><td>{{ x.customer_code }}</td><td><span class="badge {{ 'badge-loan' if x.transaction_type == 'LOAN' else 'badge-payment' }}">{{ x.transaction_type }}</span></td><td>Rs. {{ x.amount|money }}</td><td>{{ x.note or '-' }}</td><td>{{ x.created_at|sl_time }}</td><td><div class="excel-actions"><a class="icon-btn edit" title="Edit Transaction" href="{{ url_for('edit_transaction', transaction_id=x.id) }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 20h9"/><path d="M16.5 3.5a2.1 2.1 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg></a><form method="POST" action="{{ url_for('delete_transaction', transaction_id=x.id) }}" onsubmit="return confirm('Delete this transaction?');"><button class="icon-btn delete" title="Delete Transaction" type="submit"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 6h18"/><path d="M8 6V4h8v2"/><path d="M19 6l-1 14H6L5 6"/></svg></button></form></div></td></tr>{% endfor %}
      </tbody></table></div>{% else %}<div class="empty">No transaction data found.</div>{% endif %}
    </div>

    <div class="excel-sheet">
      <div class="excel-sheet-head"><div class="excel-sheet-title"><h2>Shop Items</h2><span class="data-count">{{ items|length }}</span></div><a class="icon-btn download" title="Download Shop Items Excel" href="{{ url_for('export_all_data_excel', sheet='items') }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg></a></div>
      {% if items %}<div class="excel-grid-wrap"><table class="excel-table"><thead><tr><th>#</th><th>Code</th><th>Item</th><th>Quantity</th><th>Description</th><th>Created</th><th>Actions</th></tr></thead><tbody>
      {% for x in items %}<tr><td class="row-no">{{ loop.index }}</td><td>{{ x.item_code or '-' }}</td><td>{{ x.item_name }}</td><td>{{ x.quantity }}</td><td>{{ x.description or '-' }}</td><td>{{ x.created_at|sl_time }}</td><td><div class="excel-actions"><a class="icon-btn edit" title="Edit Item" href="{{ url_for('edit_item', item_id=x.id) }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 20h9"/><path d="M16.5 3.5a2.1 2.1 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg></a><form method="POST" action="{{ url_for('delete_item', item_id=x.id) }}" onsubmit="return confirm('Delete this shop item?');"><button class="icon-btn delete" title="Delete Item" type="submit"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 6h18"/><path d="M8 6V4h8v2"/><path d="M19 6l-1 14H6L5 6"/></svg></button></form></div></td></tr>{% endfor %}
      </tbody></table></div>{% else %}<div class="empty">No shop item data found.</div>{% endif %}
    </div>

    <div class="excel-sheet">
      <div class="excel-sheet-head"><div class="excel-sheet-title"><h2>Prices</h2><span class="data-count">{{ prices|length }}</span></div><a class="icon-btn download" title="Download Prices Excel" href="{{ url_for('export_all_data_excel', sheet='prices') }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg></a></div>
      {% if prices %}<div class="excel-grid-wrap"><table class="excel-table"><thead><tr><th>#</th><th>Item</th><th>Buying Price</th><th>Selling Price</th><th>Profit / Item</th><th>Updated</th><th>Actions</th></tr></thead><tbody>
      {% for x in prices %}<tr><td class="row-no">{{ loop.index }}</td><td>{{ x.item_name }}</td><td>Rs. {{ x.buying_price|money }}</td><td>Rs. {{ x.selling_price|money }}</td><td>Rs. {{ (x.selling_price - x.buying_price)|money }}</td><td>{{ x.updated_at|sl_time }}</td><td><div class="excel-actions"><a class="icon-btn edit" title="Edit Price" href="{{ url_for('edit_price', price_id=x.id) }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 20h9"/><path d="M16.5 3.5a2.1 2.1 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg></a><form method="POST" action="{{ url_for('delete_price', price_id=x.id) }}" onsubmit="return confirm('Delete this price?');"><button class="icon-btn delete" title="Delete Price" type="submit"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M3 6h18"/><path d="M8 6V4h8v2"/><path d="M19 6l-1 14H6L5 6"/></svg></button></form></div></td></tr>{% endfor %}
      </tbody></table></div>{% else %}<div class="empty">No price data found.</div>{% endif %}
    </div>

    <div class="excel-sheet">
      <div class="excel-sheet-head"><div class="excel-sheet-title"><h2>System Users</h2><span class="data-count">{{ users|length }}</span></div><a class="icon-btn download" title="Download Users Excel" href="{{ url_for('export_all_data_excel', sheet='users') }}"><svg viewBox="0 0 24 24" fill="none" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg></a></div>
      {% if users %}<div class="excel-grid-wrap"><table class="excel-table"><thead><tr><th>#</th><th>User ID</th><th>Username</th><th>Created</th></tr></thead><tbody>
      {% for x in users %}<tr><td class="row-no">{{ loop.index }}</td><td>{{ x.id }}</td><td>{{ x.username }}</td><td>{{ x.created_at|sl_time }}</td></tr>{% endfor %}
      </tbody></table></div>{% else %}<div class="empty">No user data found.</div>{% endif %}
    </div>
    """

    return render_page(
        "All Data", "Data Store", template, active_page="database",
        customers=customers, transactions=transactions, items=items, prices=prices, users=users,
        search=search, db_name=DB_NAME, db_host=DB_HOST, db_port=DB_PORT
    )


# --------------------------- START ----------------------------

def prepare_application():
    """
    Initialize the database when the application starts.

    Gunicorn imports this module on Render, so this function is called
    outside the __main__ block as well.
    """
    try:
        if SKIP_DB_INIT_ON_STARTUP:
            initialization_message = "Startup database check skipped for fastest production boot."
        elif FAST_START and database_schema_ready():
            initialization_message = "Existing database schema detected - full initialization skipped for faster startup."
        else:
            initialize_database()
            initialization_message = "Database initialization completed."

        if os.getenv("ENSURE_PERFORMANCE_INDEXES", "true").lower() == "true":
            ensure_performance_indexes()

        print("=" * 60)
        print("AMBAAL SHOP MANAGEMENT SYSTEM")
        print(f"Database: {DB_NAME}")
        print(f"Database host: {DB_HOST}:{DB_PORT}")
        print(initialization_message)
        print("=" * 60)
    except Error as error:
        # The web server can still start and display database errors.
        # This makes configuration problems easier to diagnose on Render.
        print("=" * 60)
        print("WARNING: Could not initialize MySQL.")
        print(f"DB_HOST={DB_HOST}")
        print(f"DB_PORT={DB_PORT}")
        print(f"DB_USER={DB_USER}")
        print(f"DB_NAME={DB_NAME}")
        print(f"Error details: {error}")
        print("=" * 60)


# This runs when Gunicorn imports:
#     gunicorn ambaal_shop_system:app
prepare_application()


if __name__ == "__main__":
    # Local development:
    #     python3 ambaal_shop_system.py
    #
    # Render normally uses Gunicorn instead of this Flask development server.
    port = int(os.environ.get("PORT", "5003"))

    print("=" * 60)
    print("LOCAL DEVELOPMENT SERVER")
    print(f"Open: http://127.0.0.1:{port}")
    print(f"Admin username: {os.getenv('ADMIN_USERNAME', 'ambaal')}")
    print("=" * 60)

    app.run(
        debug=False,
        host="0.0.0.0",
        port=port
    )
